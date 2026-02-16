# main.py - 实验管理核心API

from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse, PlainTextResponse
from fastapi.encoders import jsonable_encoder
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Tuple, Set
from datetime import datetime, timezone
from enum import Enum
from copy import deepcopy
import csv
import html
import hashlib
import io
import json
import mimetypes
import re
import requests
import secrets
import shutil
import time
import uuid
import os
from urllib.parse import quote, parse_qs, urlparse, unquote
import zipfile
import xml.etree.ElementTree as ET
from email.utils import parsedate_to_datetime

try:
    from tavily import TavilyClient
except Exception:
    TavilyClient = None

app = FastAPI(title="福州理工学院AI编程实践教学平台 - 实验管理API")

# JupyterHub multi-tenant integration:
# - Internal URL is used for Hub API and proxied single-user API calls from containers.
# - Public URL is returned to browser iframes.
JUPYTERHUB_INTERNAL_URL = os.getenv("JUPYTERHUB_INTERNAL_URL", "http://jupyterhub:8000").rstrip("/")
# Prefer same-origin reverse-proxy path to avoid cross-origin cookie/WebSocket auth mismatches.
JUPYTERHUB_PUBLIC_URL = os.getenv("JUPYTERHUB_PUBLIC_URL", "/jupyter").rstrip("/")
JUPYTERHUB_API_TOKEN = os.getenv("JUPYTERHUB_API_TOKEN", "").strip()
JUPYTERHUB_REQUEST_TIMEOUT_SECONDS = float(os.getenv("JUPYTERHUB_REQUEST_TIMEOUT_SECONDS", "10"))
JUPYTERHUB_START_TIMEOUT_SECONDS = float(os.getenv("JUPYTERHUB_START_TIMEOUT_SECONDS", "60"))
# Keep user browser sessions stable for long classes.
JUPYTERHUB_USER_TOKEN_EXPIRES_SECONDS = int(os.getenv("JUPYTERHUB_USER_TOKEN_EXPIRES_SECONDS", "43200"))
JUPYTER_WORKSPACE_UI = str(os.getenv("JUPYTER_WORKSPACE_UI", "lab") or "").strip().lower()
TAVILY_API_KEY = os.getenv("TAVILY_API_KEY", "").strip()

def _parse_account_list(raw: str) -> List[str]:
    parts = [item.strip() for item in str(raw or "").split(",")]
    return [item for item in parts if item]


# 教师账号列表
TEACHER_ACCOUNTS = _parse_account_list(
    os.getenv("TEACHER_ACCOUNTS", "teacher_001,teacher_002,teacher_003,teacher_004,teacher_005")
)
ADMIN_ACCOUNTS = _parse_account_list(os.getenv("ADMIN_ACCOUNTS", "admin"))

def is_teacher(username: str) -> bool:
    """判断用户是否为教师"""
    normalized = _normalize_text(username)
    if not normalized:
        return False
    return normalized in TEACHER_ACCOUNTS or normalized in teachers_db


def is_admin(username: str) -> bool:
    """判断用户是否为管理员"""
    return username in ADMIN_ACCOUNTS


DEFAULT_PASSWORD = "123456"
UPLOAD_DIR = "/app/uploads"
USER_REGISTRY_FILE = os.path.join(UPLOAD_DIR, "user_registry.json")
RESOURCE_REGISTRY_FILE = os.path.join(UPLOAD_DIR, "resource_registry.json")
EXPERIMENT_REGISTRY_FILE = os.path.join(UPLOAD_DIR, "experiment_registry.json")
COURSE_REGISTRY_FILE = os.path.join(UPLOAD_DIR, "course_registry.json")
ATTACHMENT_REGISTRY_FILE = os.path.join(UPLOAD_DIR, "attachment_registry.json")
AI_SHARED_CONFIG_FILE = os.path.join(UPLOAD_DIR, "ai_shared_config.json")
AI_CHAT_HISTORY_FILE = os.path.join(UPLOAD_DIR, "ai_chat_history.json")
RESOURCE_POLICY_FILE = os.path.join(UPLOAD_DIR, "user_resource_policy.json")
OPERATION_LOG_FILE = os.path.join(UPLOAD_DIR, "operation_log.json")
SEED_MARKER_FILE = os.path.join(UPLOAD_DIR, ".seed_defaults_v1")  # legacy filename (kept for backward compat)
TEXT_PREVIEW_CHAR_LIMIT = 20000
ALLOWED_RESOURCE_EXTENSIONS = {
    ".pdf",
    ".doc",
    ".docx",
    ".md",
    ".markdown",
    ".txt",
    ".csv",
    ".json",
    ".ppt",
    ".pptx",
    ".xls",
    ".xlsx",
}
TEMPLATE_HEADERS = ["学号", "姓名", "班级", "单位名称", "手机号", "入学年级"]
LEGACY_TEMPLATE_HEADERS = TEMPLATE_HEADERS[:5]
DEFAULT_ADMISSION_YEAR_OPTIONS = ["2020", "2021", "2022", "2023", "2024", "2025", "2026", "2027", "2028"]
CLASS_TEMPLATE_HEADERS = ["入学年级", "专业", "班级"]
DEFAULT_AI_SHARED_CONFIG = {
    "api_key": "",
    "tavily_api_key": "",
    "chat_model": "deepseek-chat",
    "reasoner_model": "deepseek-reasoner",
    "base_url": "https://api.deepseek.com",
    "system_prompt": "你是福州理工学院AI编程实践教学平台小助手。请使用简洁、准确、教学友好的中文回答。"
}
AI_RESPONSE_STYLE_RULES = (
    "回答规则：先给结论，再给关键依据或步骤；"
    "代码问题优先给最小可运行示例；"
    "避免空话和套话，不要编造事实；"
    "不确定时明确写“我不确定/需要进一步检索确认”。"
)
DEFAULT_RESOURCE_ROLE_LIMITS = {
    "student": {"cpu_limit": 2.0, "memory_limit": "8G", "storage_limit": "2G"},
    "teacher": {"cpu_limit": 2.0, "memory_limit": "8G", "storage_limit": "2G"},
    "admin": {"cpu_limit": 4.0, "memory_limit": "8G", "storage_limit": "20G"},
}
DEFAULT_SERVER_RESOURCE_BUDGET = {
    "max_total_cpu": 64.0,
    "max_total_memory": "128G",
    "max_total_storage": "1T",
    "enforce_budget": False,
}
MAX_OPERATION_LOG_ITEMS = 5000
AI_CHAT_HISTORY_MAX_MESSAGES = max(20, int(os.getenv("AI_CHAT_HISTORY_MAX_MESSAGES", "240")))
AI_CHAT_HISTORY_MAX_MESSAGE_CHARS = max(1000, int(os.getenv("AI_CHAT_HISTORY_MAX_MESSAGE_CHARS", "12000")))
AI_CONTEXT_MAX_HISTORY_MESSAGES = max(10, int(os.getenv("AI_CONTEXT_MAX_HISTORY_MESSAGES", "80")))
AI_CONTEXT_MAX_TOTAL_CHARS = max(4000, int(os.getenv("AI_CONTEXT_MAX_TOTAL_CHARS", "48000")))
AI_SESSION_TTL_SECONDS = max(900, int(os.getenv("AI_SESSION_TTL_SECONDS", "43200")))
AI_SESSION_MAX_TOKENS = max(100, int(os.getenv("AI_SESSION_MAX_TOKENS", "5000")))
AI_WEB_SEARCH_CACHE_TTL_SECONDS = max(60, int(os.getenv("AI_WEB_SEARCH_CACHE_TTL_SECONDS", "3600")))
AI_WEB_SEARCH_CACHE_MAX_ITEMS = max(50, int(os.getenv("AI_WEB_SEARCH_CACHE_MAX_ITEMS", "1000")))
PASSWORD_HASH_PATTERN = re.compile(r"^[0-9a-f]{64}$")

os.makedirs(UPLOAD_DIR, exist_ok=True)


def _read_seed_marker() -> Tuple[int, dict]:
    """Return (version, payload). Version 0 means not seeded.

    Legacy marker content was a plain timestamp string; treat that as version 1.
    """
    if not os.path.exists(SEED_MARKER_FILE):
        return 0, {}

    try:
        with open(SEED_MARKER_FILE, "r", encoding="utf-8") as file_obj:
            raw = (file_obj.read() or "").strip()
    except OSError as exc:
        print(f"Failed to read seed marker: {exc}")
        return 1, {}

    if not raw:
        return 1, {}

    try:
        payload = json.loads(raw)
    except json.JSONDecodeError:
        return 1, {"legacy": True, "raw": raw}

    if isinstance(payload, dict):
        version = payload.get("version", 1)
        try:
            version = int(version)
        except (TypeError, ValueError):
            version = 1
        return version, payload

    return 1, {}


def _write_seed_marker(version: int, payload: Optional[dict] = None):
    data = {"version": int(version), "updated_at": datetime.now().isoformat()}
    if payload:
        data.update(payload)

    tmp_path = f"{SEED_MARKER_FILE}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as file_obj:
        json.dump(data, file_obj, ensure_ascii=False, indent=2)
    os.replace(tmp_path, SEED_MARKER_FILE)


def _empty_notebook_json() -> dict:
    return {"cells": [], "metadata": {}, "nbformat": 4, "nbformat_minor": 5}


def _jupyterhub_enabled() -> bool:
    return bool(JUPYTERHUB_INTERNAL_URL and JUPYTERHUB_API_TOKEN)


def _hub_api_url(path: str) -> str:
    normalized = (path or "").strip()
    if not normalized.startswith("/"):
        normalized = f"/{normalized}"
    return f"{JUPYTERHUB_INTERNAL_URL}{normalized}"


def _hub_headers(extra: Optional[dict] = None) -> dict:
    headers = {"Authorization": f"token {JUPYTERHUB_API_TOKEN}"} if JUPYTERHUB_API_TOKEN else {}
    if extra:
        headers.update(extra)
    return headers


def _hub_request(method: str, path: str, **kwargs) -> requests.Response:
    headers = kwargs.pop("headers", None)
    timeout = kwargs.pop("timeout", JUPYTERHUB_REQUEST_TIMEOUT_SECONDS)
    allow_redirects = kwargs.pop("allow_redirects", False)
    return requests.request(
        method,
        _hub_api_url(path),
        headers=_hub_headers(headers),
        timeout=timeout,
        allow_redirects=allow_redirects,
        **kwargs,
    )


def _ensure_hub_user_exists(username: str) -> bool:
    user = _normalize_text(username)
    if not user or not _jupyterhub_enabled():
        return False

    try:
        resp = _hub_request("GET", f"/hub/api/users/{quote(user)}")
        if resp.status_code == 200:
            return True
        if resp.status_code != 404:
            print(f"JupyterHub user lookup failed ({resp.status_code}): {resp.text[:200]}")
            return False
    except requests.RequestException as exc:
        print(f"JupyterHub user lookup error: {exc}")
        return False

    try:
        create_resp = _hub_request("POST", f"/hub/api/users/{quote(user)}")
        if create_resp.status_code in {201, 200}:
            return True
        if create_resp.status_code == 409:
            return True
        print(f"JupyterHub user create failed ({create_resp.status_code}): {create_resp.text[:200]}")
    except requests.RequestException as exc:
        print(f"JupyterHub user create error: {exc}")

    return False


def _ensure_user_server_running(username: str) -> bool:
    user = _normalize_text(username)
    if not user or not _jupyterhub_enabled():
        return False

    if not _ensure_hub_user_exists(user):
        return False

    try:
        resp = _hub_request("POST", f"/hub/api/users/{quote(user)}/server")
        # 201/202: started, 409: already running
        if resp.status_code not in {201, 202, 409}:
            # Some JupyterHub versions return 400 with a message when the server is already running.
            if resp.status_code == 400:
                try:
                    payload = resp.json() or {}
                except ValueError:
                    payload = {}
                message = str(payload.get("message") or payload.get("detail") or resp.text or "")
                if "already running" in message.lower():
                    resp = None
                else:
                    print(f"JupyterHub spawn failed ({resp.status_code}): {message[:200]}")
                    return False
            else:
                print(f"JupyterHub spawn failed ({resp.status_code}): {resp.text[:200]}")
                return False
    except requests.RequestException as exc:
        print(f"JupyterHub spawn error: {exc}")
        return False

    deadline = time.time() + max(5.0, JUPYTERHUB_START_TIMEOUT_SECONDS)
    while time.time() < deadline:
        try:
            status = _hub_request("GET", f"/hub/api/users/{quote(user)}")
            if status.status_code != 200:
                time.sleep(1)
                continue
            payload = status.json()
            pending = payload.get("pending")
            if pending:
                time.sleep(1)
                continue

            server_field = payload.get("server")
            if server_field:
                return True

            servers = payload.get("servers")
            if isinstance(servers, dict):
                for srv in servers.values():
                    if srv:
                        if isinstance(srv, dict) and srv.get("pending"):
                            continue
                        return True
        except Exception:
            pass
        time.sleep(1)

    return False


def _wait_user_server_state(username: str, expect_running: bool, timeout_seconds: Optional[float] = None) -> bool:
    user = _normalize_text(username)
    if not user:
        return False

    deadline = time.time() + max(5.0, float(timeout_seconds or JUPYTERHUB_START_TIMEOUT_SECONDS))
    while time.time() < deadline:
        try:
            status = _hub_request("GET", f"/hub/api/users/{quote(user)}")
            if status.status_code == 404:
                return not expect_running
            if status.status_code != 200:
                time.sleep(1)
                continue

            payload = status.json() or {}
            state = _extract_server_state(payload)
            is_running = bool(state.get("server_running"))
            is_pending = bool(state.get("server_pending"))

            if expect_running and is_running:
                return True
            if not expect_running and (not is_running and not is_pending):
                return True
        except Exception:
            pass
        time.sleep(1)
    return False


def _stop_user_server(username: str) -> bool:
    user = _normalize_text(username)
    if not user or not _jupyterhub_enabled():
        return False

    try:
        user_resp = _hub_request("GET", f"/hub/api/users/{quote(user)}")
        if user_resp.status_code == 404:
            return True
        if user_resp.status_code != 200:
            print(f"JupyterHub user lookup failed ({user_resp.status_code}): {user_resp.text[:200]}")
            return False
    except requests.RequestException as exc:
        print(f"JupyterHub user lookup error: {exc}")
        return False

    try:
        stop_resp = _hub_request("DELETE", f"/hub/api/users/{quote(user)}/server")
        if stop_resp.status_code not in {202, 204, 404}:
            if stop_resp.status_code == 400:
                try:
                    payload = stop_resp.json() or {}
                except ValueError:
                    payload = {}
                message = str(payload.get("message") or payload.get("detail") or stop_resp.text or "")
                normalized_message = message.lower()
                if "not running" not in normalized_message and "no such server" not in normalized_message:
                    print(f"JupyterHub stop failed ({stop_resp.status_code}): {message[:200]}")
                    return False
            else:
                print(f"JupyterHub stop failed ({stop_resp.status_code}): {stop_resp.text[:200]}")
                return False
    except requests.RequestException as exc:
        print(f"JupyterHub stop error: {exc}")
        return False

    return _wait_user_server_state(user, expect_running=False)


def _create_short_lived_user_token(
    username: str,
    expires_in: int = JUPYTERHUB_USER_TOKEN_EXPIRES_SECONDS,
) -> Optional[str]:
    user = _normalize_text(username)
    if not user or not _jupyterhub_enabled():
        return None
    if not _ensure_hub_user_exists(user):
        return None

    try:
        resp = _hub_request(
            "POST",
            f"/hub/api/users/{quote(user)}/tokens",
            # Restrict scope to only allow accessing this user's server via the proxy.
            json={
                "note": "training-platform",
                "expires_in": int(expires_in),
                "scopes": [f"access:servers!user={user}"],
            },
        )
        if resp.status_code not in {200, 201}:
            print(f"JupyterHub create token failed ({resp.status_code}): {resp.text[:200]}")
            return None
        return (resp.json() or {}).get("token")
    except requests.RequestException as exc:
        print(f"JupyterHub create token error: {exc}")
        return None


def _user_contents_url(username: str, path: str) -> str:
    user = _normalize_text(username)
    normalized_path = (path or "").lstrip("/")
    encoded_path = quote(normalized_path, safe="/")
    return f"{JUPYTERHUB_INTERNAL_URL}/user/{quote(user)}/api/contents/{encoded_path}"


def _user_contents_request(username: str, token: str, method: str, path: str, **kwargs) -> requests.Response:
    headers = kwargs.pop("headers", None) or {}
    headers = {**headers, "Authorization": f"token {token}"}
    timeout = kwargs.pop("timeout", JUPYTERHUB_REQUEST_TIMEOUT_SECONDS)
    allow_redirects = kwargs.pop("allow_redirects", False)
    return requests.request(
        method,
        _user_contents_url(username, path),
        headers=headers,
        timeout=timeout,
        allow_redirects=allow_redirects,
        **kwargs,
    )


def _append_token(url: str, token: Optional[str]) -> str:
    if not token:
        return url
    separator = "&" if "?" in url else "?"
    return f"{url}{separator}token={quote(token)}"


def _build_user_lab_url(username: str, path: Optional[str] = None, token: Optional[str] = None) -> str:
    user = _normalize_text(username)
    if not user:
        return ""

    workspace_ui = JUPYTER_WORKSPACE_UI if JUPYTER_WORKSPACE_UI in {"notebook", "lab"} else "lab"

    if workspace_ui == "notebook":
        if path:
            encoded_path = quote(path, safe="/")
            base = f"{JUPYTERHUB_PUBLIC_URL}/user/{quote(user)}/notebooks/{encoded_path}"
        else:
            base = f"{JUPYTERHUB_PUBLIC_URL}/user/{quote(user)}/tree"
    else:
        if path:
            encoded_path = quote(path, safe="/")
            base = f"{JUPYTERHUB_PUBLIC_URL}/user/{quote(user)}/lab/tree/{encoded_path}"
        else:
            base = f"{JUPYTERHUB_PUBLIC_URL}/user/{quote(user)}/lab"

    return _append_token(base, token)

# CORS配置
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== 数据模型 ====================

class DifficultyLevel(str, Enum):
    BEGINNER = "初级"
    INTERMEDIATE = "中级"
    ADVANCED = "高级"

class ExperimentStatus(str, Enum):
    NOT_STARTED = "未开始"
    IN_PROGRESS = "进行中"
    SUBMITTED = "已提交"
    GRADED = "已评分"


class PublishScope(str, Enum):
    ALL = "all"
    CLASS = "class"
    STUDENT = "student"

class Experiment(BaseModel):
    id: str = None
    course_id: Optional[str] = None
    course_name: str = "Python程序设计"
    title: str
    description: Optional[str] = None
    difficulty: DifficultyLevel = DifficultyLevel.BEGINNER
    tags: List[str] = []
    notebook_path: Optional[str] = None
    resources: dict = {"cpu": 1.0, "memory": "2G", "storage": "1G"}
    deadline: Optional[datetime] = None
    created_at: datetime = None
    created_by: str
    published: bool = True  # 是否发布给学生
    publish_scope: PublishScope = PublishScope.ALL
    target_class_names: List[str] = Field(default_factory=list)
    target_student_ids: List[str] = Field(default_factory=list)
    
    class Config:
        json_encoders = {
            datetime: lambda v: v.isoformat()
        }

class StudentExperiment(BaseModel):
    id: str = None
    experiment_id: str
    student_id: str
    status: ExperimentStatus = ExperimentStatus.NOT_STARTED
    start_time: Optional[datetime] = None
    submit_time: Optional[datetime] = None
    notebook_content: Optional[str] = None
    score: Optional[float] = None
    ai_feedback: Optional[str] = None
    teacher_comment: Optional[str] = None

class SubmitExperimentRequest(BaseModel):
    notebook_content: str


class LoginRequest(BaseModel):
    username: str
    password: str


class StudentPasswordChangeRequest(BaseModel):
    student_id: str
    old_password: str
    new_password: str


class StudentSecurityQuestionUpdateRequest(BaseModel):
    student_id: str
    security_question: str
    security_answer: str


class TeacherPasswordChangeRequest(BaseModel):
    teacher_username: str
    old_password: str
    new_password: str


class TeacherSecurityQuestionUpdateRequest(BaseModel):
    teacher_username: str
    security_question: str
    security_answer: str


class ForgotPasswordResetRequest(BaseModel):
    username: str
    security_answer: str
    new_password: str


class CourseCreateRequest(BaseModel):
    name: str
    description: Optional[str] = ""
    teacher_username: str


class CourseUpdateRequest(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    teacher_username: str


class ClassCreateRequest(BaseModel):
    name: str
    teacher_username: str


class TeacherCreateRequest(BaseModel):
    admin_username: str
    username: str
    real_name: Optional[str] = ""


class CourseRecord(BaseModel):
    id: str
    name: str
    description: Optional[str] = ""
    created_by: str
    created_at: datetime
    updated_at: datetime


class ClassRecord(BaseModel):
    id: str
    name: str
    created_by: str
    created_at: datetime


class TeacherRecord(BaseModel):
    username: str
    real_name: str = ""
    created_by: str
    created_at: datetime


class StudentRecord(BaseModel):
    student_id: str
    username: str
    real_name: str
    class_name: str
    admission_year: str = ""
    organization: str
    phone: str
    role: str = "student"
    created_by: str = ""
    password_hash: str
    security_question: str = ""
    security_answer_hash: str = ""
    created_at: datetime
    updated_at: datetime


class PDFAnnotation(BaseModel):
    id: str
    teacher_username: str
    content: str
    created_at: datetime


class StudentSubmissionPDF(BaseModel):
    id: str
    student_exp_id: str
    experiment_id: str
    student_id: str
    filename: str
    file_path: str
    content_type: str
    size: int
    created_at: datetime
    viewed: bool = False
    viewed_at: Optional[datetime] = None
    viewed_by: Optional[str] = None
    reviewed: bool = False
    reviewed_at: Optional[datetime] = None
    reviewed_by: Optional[str] = None
    annotations: List[PDFAnnotation] = Field(default_factory=list)


class PDFAnnotationCreateRequest(BaseModel):
    teacher_username: str
    content: str


class ResourceFile(BaseModel):
    id: str
    filename: str
    file_path: str
    file_type: str
    content_type: str
    size: int
    created_at: datetime
    created_by: str


class ResourceQuotaUpdateRequest(BaseModel):
    admin_username: str
    cpu_limit: float = Field(..., gt=0, le=128)
    memory_limit: str = Field(..., min_length=1, max_length=32)
    storage_limit: str = Field(..., min_length=1, max_length=32)
    note: Optional[str] = ""


class ResourceBudgetUpdateRequest(BaseModel):
    admin_username: str
    max_total_cpu: float = Field(..., gt=0, le=1024)
    max_total_memory: str = Field(..., min_length=1, max_length=32)
    max_total_storage: str = Field(..., min_length=1, max_length=32)
    enforce_budget: bool = False


class OperationLogEntry(BaseModel):
    id: str
    operator: str
    action: str
    target: str
    detail: str = ""
    success: bool = True
    created_at: datetime


# ==================== 模拟数据库 ====================
# 生产环境应使用 PostgreSQL + SQLAlchemy

experiments_db = {}
courses_db: Dict[str, CourseRecord] = {}
student_experiments_db = {}
classes_db: Dict[str, ClassRecord] = {}
teachers_db: Dict[str, TeacherRecord] = {}
students_db: Dict[str, StudentRecord] = {}
teacher_account_password_hashes_db: Dict[str, str] = {}
account_security_questions_db: Dict[str, Dict[str, str]] = {}
submission_pdfs_db: Dict[str, StudentSubmissionPDF] = {}
resource_files_db: Dict[str, ResourceFile] = {}
ai_shared_config_db: Dict[str, str] = dict(DEFAULT_AI_SHARED_CONFIG)
ai_chat_history_db: Dict[str, List[Dict[str, str]]] = {}
ai_session_tokens_db: Dict[str, Dict[str, object]] = {}
ai_web_search_cache_db: Dict[str, Dict[str, object]] = {}
resource_policy_db: Dict[str, dict] = {}
operation_logs_db: List[OperationLogEntry] = []


def _hash_password(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def _default_password_hash() -> str:
    return _hash_password(DEFAULT_PASSWORD)


def _get_account_password_hash(username: str) -> str:
    normalized = _normalize_text(username)
    if not normalized:
        return _default_password_hash()

    saved_hash = _normalize_text(teacher_account_password_hashes_db.get(normalized)).lower()
    if saved_hash and PASSWORD_HASH_PATTERN.fullmatch(saved_hash):
        return saved_hash
    return _default_password_hash()


def _verify_account_password(username: str, password: str) -> bool:
    return _hash_password(password or "") == _get_account_password_hash(username)


def _normalize_security_question(question: str) -> str:
    return _normalize_text(question)[:120]


def _normalize_security_answer(answer: str) -> str:
    return _normalize_text(answer).lower()[:200]


def _hash_security_answer(answer: str) -> str:
    normalized = _normalize_security_answer(answer)
    return _hash_password(normalized)


def _verify_security_answer(stored_hash: str, provided_answer: str) -> bool:
    normalized_hash = _normalize_text(stored_hash).lower()
    if not PASSWORD_HASH_PATTERN.fullmatch(normalized_hash):
        return False
    return normalized_hash == _hash_security_answer(provided_answer)


def _normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cleanup_ai_sessions(now_ts: Optional[float] = None):
    now_value = float(now_ts if now_ts is not None else time.time())

    expired_tokens = [
        token
        for token, item in ai_session_tokens_db.items()
        if float(item.get("expires_at") or 0.0) <= now_value
    ]
    for token in expired_tokens:
        ai_session_tokens_db.pop(token, None)

    if len(ai_session_tokens_db) <= AI_SESSION_MAX_TOKENS:
        return

    sorted_items = sorted(
        ai_session_tokens_db.items(),
        key=lambda pair: float((pair[1] or {}).get("expires_at") or 0.0),
    )
    overflow = len(sorted_items) - AI_SESSION_MAX_TOKENS
    for token, _ in sorted_items[:overflow]:
        ai_session_tokens_db.pop(token, None)


def _create_ai_session_token(username: str) -> str:
    normalized_user = _normalize_text(username)
    if not normalized_user:
        return ""

    now_ts = time.time()
    _cleanup_ai_sessions(now_ts)

    token = secrets.token_urlsafe(36)
    ai_session_tokens_db[token] = {
        "username": normalized_user,
        "expires_at": now_ts + AI_SESSION_TTL_SECONDS,
    }
    return token


def _resolve_ai_session_user(token: str) -> str:
    normalized_token = _normalize_text(token)
    if not normalized_token:
        return ""

    now_ts = time.time()
    _cleanup_ai_sessions(now_ts)

    session_item = ai_session_tokens_db.get(normalized_token) or {}
    username = _normalize_text(session_item.get("username"))
    expires_at = float(session_item.get("expires_at") or 0.0)
    if not username or expires_at <= now_ts:
        ai_session_tokens_db.pop(normalized_token, None)
        return ""

    # Sliding window refresh for active users.
    session_item["expires_at"] = now_ts + AI_SESSION_TTL_SECONDS
    ai_session_tokens_db[normalized_token] = session_item
    return username


def _require_ai_session(
    request: Request,
    *,
    expected_username: Optional[str] = None,
    allow_admin_override: bool = True,
) -> str:
    token = _normalize_text(request.headers.get("X-AI-Session-Token"))
    if not token:
        raise HTTPException(status_code=401, detail="AI会话不存在或已过期，请重新登录")

    actor = _resolve_ai_session_user(token)
    if not actor:
        raise HTTPException(status_code=401, detail="AI会话不存在或已过期，请重新登录")

    normalized_expected = _normalize_text(expected_username)
    if normalized_expected and actor != normalized_expected:
        if not (allow_admin_override and is_admin(actor)):
            raise HTTPException(status_code=403, detail="无权访问该用户AI数据")

    return actor


def _normalize_admission_year(value) -> str:
    raw = _normalize_text(value)
    if not raw:
        return ""
    digits = "".join(ch for ch in raw if ch.isdigit())
    if len(digits) == 4 and digits.startswith("20"):
        return digits
    if len(digits) == 2:
        return f"20{digits}"
    return ""


def _all_teacher_accounts() -> List[str]:
    merged = set()
    for username in TEACHER_ACCOUNTS:
        normalized = _normalize_text(username)
        if normalized:
            merged.add(normalized)
    for username in teachers_db.keys():
        normalized = _normalize_text(username)
        if normalized:
            merged.add(normalized)
    return sorted(merged)


def _infer_user_role(username: str) -> str:
    normalized = _normalize_text(username)
    if is_admin(normalized):
        return "admin"
    if is_teacher(normalized):
        return "teacher"
    return "student"


def _is_student_progress_record(student_id: str) -> bool:
    """Only keep student-origin records in teacher-facing progress/review views."""
    normalized = _normalize_text(student_id)
    if not normalized:
        return False
    if is_admin(normalized) or is_teacher(normalized):
        return False
    return True


def _ensure_admin(admin_username: str):
    normalized = _normalize_text(admin_username)
    if not is_admin(normalized):
        raise HTTPException(status_code=403, detail="权限不足，需要管理员账号")


_SIZE_LIMIT_PATTERN = re.compile(r"^\s*(\d+(?:\.\d+)?)\s*([kmgt]?b?)?\s*$", re.IGNORECASE)
_SIZE_FACTORS = {
    "B": 1,
    "K": 1024,
    "M": 1024 ** 2,
    "G": 1024 ** 3,
    "T": 1024 ** 4,
}


def _default_size_unit(default_value: str) -> str:
    match = _SIZE_LIMIT_PATTERN.match(_normalize_text(default_value))
    if not match:
        return "B"
    unit_raw = (match.group(2) or "").upper()
    if unit_raw in {"K", "KB"}:
        return "K"
    if unit_raw in {"M", "MB"}:
        return "M"
    if unit_raw in {"G", "GB"}:
        return "G"
    if unit_raw in {"T", "TB"}:
        return "T"
    return "B"


def _normalize_size_limit(value, default_value: str) -> str:
    raw = _normalize_text(value)
    if not raw:
        return default_value

    match = _SIZE_LIMIT_PATTERN.match(raw)
    if not match:
        raise HTTPException(status_code=400, detail=f"资源大小格式无效: {raw}")

    number = float(match.group(1))
    if number <= 0:
        raise HTTPException(status_code=400, detail=f"资源大小必须大于 0: {raw}")

    default_unit = _default_size_unit(default_value)
    unit_raw = (match.group(2) or "").upper()
    if unit_raw == "":
        unit = default_unit
    elif unit_raw == "B":
        # Backward compatibility: older UI values like "8" were normalized into "8B".
        unit = default_unit if default_unit != "B" else "B"
    elif unit_raw in {"K", "KB"}:
        unit = "K"
    elif unit_raw in {"M", "MB"}:
        unit = "M"
    elif unit_raw in {"G", "GB"}:
        unit = "G"
    elif unit_raw in {"T", "TB"}:
        unit = "T"
    else:
        raise HTTPException(status_code=400, detail=f"资源大小单位无效: {raw}")

    if number.is_integer():
        number_text = str(int(number))
    else:
        number_text = str(round(number, 3)).rstrip("0").rstrip(".")
    if unit == "B":
        return number_text
    return f"{number_text}{unit}"


def _size_to_bytes(value: str) -> int:
    normalized = _normalize_size_limit(value, "0B")
    match = _SIZE_LIMIT_PATTERN.match(normalized)
    if not match:
        return 0
    number = float(match.group(1))
    unit_raw = (match.group(2) or "B").upper()
    if unit_raw in {"", "B"}:
        unit = "B"
    elif unit_raw in {"K", "KB"}:
        unit = "K"
    elif unit_raw in {"M", "MB"}:
        unit = "M"
    elif unit_raw in {"G", "GB"}:
        unit = "G"
    elif unit_raw in {"T", "TB"}:
        unit = "T"
    else:
        unit = "B"
    return int(number * _SIZE_FACTORS[unit])


def _default_resource_policy_payload() -> dict:
    now_iso = datetime.now().isoformat()
    payload = {
        "defaults": deepcopy(DEFAULT_RESOURCE_ROLE_LIMITS),
        "budget": {
            **deepcopy(DEFAULT_SERVER_RESOURCE_BUDGET),
            "updated_by": "system",
            "updated_at": now_iso,
        },
        "overrides": {},
    }
    return payload


def _normalize_resource_quota(raw: Optional[dict], role: str) -> dict:
    role_key = role if role in DEFAULT_RESOURCE_ROLE_LIMITS else "student"
    default_quota = DEFAULT_RESOURCE_ROLE_LIMITS[role_key]
    source = raw or {}

    cpu_raw = source.get("cpu_limit", default_quota["cpu_limit"])
    try:
        cpu_limit = float(cpu_raw)
    except (TypeError, ValueError):
        cpu_limit = float(default_quota["cpu_limit"])
    cpu_limit = round(max(0.1, min(cpu_limit, 128.0)), 3)

    memory_limit = _normalize_size_limit(source.get("memory_limit", default_quota["memory_limit"]), default_quota["memory_limit"])
    storage_limit = _normalize_size_limit(source.get("storage_limit", default_quota["storage_limit"]), default_quota["storage_limit"])
    return {
        "cpu_limit": cpu_limit,
        "memory_limit": memory_limit,
        "storage_limit": storage_limit,
    }


def _normalize_resource_budget(raw: Optional[dict]) -> dict:
    source = raw or {}
    default_budget = DEFAULT_SERVER_RESOURCE_BUDGET
    try:
        max_total_cpu = float(source.get("max_total_cpu", default_budget["max_total_cpu"]))
    except (TypeError, ValueError):
        max_total_cpu = float(default_budget["max_total_cpu"])
    max_total_cpu = round(max(0.1, min(max_total_cpu, 1024.0)), 3)

    max_total_memory = _normalize_size_limit(
        source.get("max_total_memory", default_budget["max_total_memory"]),
        default_budget["max_total_memory"],
    )
    max_total_storage = _normalize_size_limit(
        source.get("max_total_storage", default_budget["max_total_storage"]),
        default_budget["max_total_storage"],
    )
    enforce_budget = bool(source.get("enforce_budget", default_budget["enforce_budget"]))
    updated_by = _normalize_text(source.get("updated_by")) or "system"
    updated_at = _normalize_text(source.get("updated_at")) or datetime.now().isoformat()
    return {
        "max_total_cpu": max_total_cpu,
        "max_total_memory": max_total_memory,
        "max_total_storage": max_total_storage,
        "enforce_budget": enforce_budget,
        "updated_by": updated_by,
        "updated_at": updated_at,
    }


def _save_resource_policy():
    tmp_path = f"{RESOURCE_POLICY_FILE}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as file_obj:
        json.dump(resource_policy_db, file_obj, ensure_ascii=False, indent=2)
    os.replace(tmp_path, RESOURCE_POLICY_FILE)


def _load_resource_policy():
    resource_policy_db.clear()
    resource_policy_db.update(_default_resource_policy_payload())
    if not os.path.exists(RESOURCE_POLICY_FILE):
        return

    try:
        with open(RESOURCE_POLICY_FILE, "r", encoding="utf-8") as file_obj:
            payload = json.load(file_obj) or {}
    except Exception as exc:
        print(f"Failed to load resource policy: {exc}")
        return

    defaults = {}
    raw_defaults = payload.get("defaults", {}) if isinstance(payload, dict) else {}
    for role in DEFAULT_RESOURCE_ROLE_LIMITS:
        defaults[role] = _normalize_resource_quota(raw_defaults.get(role), role)

    budget = _normalize_resource_budget(payload.get("budget", {}))
    overrides = {}
    raw_overrides = payload.get("overrides", {}) if isinstance(payload, dict) else {}
    if isinstance(raw_overrides, dict):
        for username, quota in raw_overrides.items():
            normalized_username = _normalize_text(username)
            if not normalized_username:
                continue
            role = _infer_user_role(normalized_username)
            normalized_quota = _normalize_resource_quota(quota, role)
            normalized_quota["updated_by"] = _normalize_text((quota or {}).get("updated_by")) or "system"
            normalized_quota["updated_at"] = _normalize_text((quota or {}).get("updated_at")) or datetime.now().isoformat()
            normalized_quota["note"] = _normalize_text((quota or {}).get("note"))[:200]
            overrides[normalized_username] = normalized_quota

    resource_policy_db.update({
        "defaults": defaults,
        "budget": budget,
        "overrides": overrides,
    })


def _operation_log_to_dict(record: OperationLogEntry) -> dict:
    payload = record.dict()
    payload["created_at"] = record.created_at.isoformat()
    return payload


def _save_operation_logs():
    payload = {"items": [_operation_log_to_dict(item) for item in operation_logs_db]}
    tmp_path = f"{OPERATION_LOG_FILE}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as file_obj:
        json.dump(payload, file_obj, ensure_ascii=False, indent=2)
    os.replace(tmp_path, OPERATION_LOG_FILE)


def _load_operation_logs():
    operation_logs_db.clear()
    if not os.path.exists(OPERATION_LOG_FILE):
        return

    try:
        with open(OPERATION_LOG_FILE, "r", encoding="utf-8") as file_obj:
            payload = json.load(file_obj) or {}
    except Exception as exc:
        print(f"Failed to load operation logs: {exc}")
        return

    for item in payload.get("items", []):
        try:
            operation_logs_db.append(OperationLogEntry(**item))
        except Exception as exc:
            print(f"Invalid operation log skipped: {exc}")


def _append_operation_log(operator: str, action: str, target: str, detail: str = "", success: bool = True):
    entry = OperationLogEntry(
        id=str(uuid.uuid4()),
        operator=_normalize_text(operator) or "unknown",
        action=_normalize_text(action) or "unknown",
        target=_normalize_text(target) or "-",
        detail=_normalize_text(detail)[:800],
        success=bool(success),
        created_at=datetime.now(),
    )
    operation_logs_db.append(entry)
    if len(operation_logs_db) > MAX_OPERATION_LOG_ITEMS:
        del operation_logs_db[: len(operation_logs_db) - MAX_OPERATION_LOG_ITEMS]
    _save_operation_logs()


def _managed_users() -> List[dict]:
    users: List[dict] = []
    seen = set()

    for username in ADMIN_ACCOUNTS:
        normalized = _normalize_text(username)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        users.append({
            "username": normalized,
            "role": "admin",
            "real_name": normalized,
            "student_id": "",
            "class_name": "",
            "organization": "",
        })

    for username in _all_teacher_accounts():
        normalized = _normalize_text(username)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        teacher = teachers_db.get(normalized)
        users.append({
            "username": normalized,
            "role": "teacher",
            "real_name": _normalize_text(getattr(teacher, "real_name", "")) or normalized,
            "student_id": "",
            "class_name": "",
            "organization": "",
        })

    for student in students_db.values():
        normalized = _normalize_text(student.username or student.student_id)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        users.append({
            "username": normalized,
            "role": "student",
            "real_name": student.real_name,
            "student_id": student.student_id,
            "class_name": student.class_name,
            "organization": student.organization,
        })

    role_order = {"admin": 0, "teacher": 1, "student": 2}
    users.sort(key=lambda item: (role_order.get(item["role"], 9), item["username"]))
    return users


def _get_role_default_quota(role: str) -> dict:
    defaults = resource_policy_db.get("defaults", {})
    role_key = role if role in DEFAULT_RESOURCE_ROLE_LIMITS else "student"
    return _normalize_resource_quota(defaults.get(role_key), role_key)


def _get_effective_user_quota(username: str, role: str, overrides: Optional[dict] = None) -> Tuple[dict, str, dict]:
    normalized_username = _normalize_text(username)
    default_quota = _get_role_default_quota(role)
    override_source = overrides if isinstance(overrides, dict) else resource_policy_db.get("overrides", {})
    override = override_source.get(normalized_username)
    if isinstance(override, dict):
        quota = _normalize_resource_quota(override, role)
        meta = {
            "updated_by": _normalize_text(override.get("updated_by")) or "unknown",
            "updated_at": _normalize_text(override.get("updated_at")) or "",
            "note": _normalize_text(override.get("note")),
        }
        return quota, "custom", meta
    return default_quota, "default", {"updated_by": "system", "updated_at": "", "note": ""}


def _hub_user_state_map() -> Dict[str, dict]:
    if not _jupyterhub_enabled():
        return {}
    try:
        resp = _hub_request("GET", "/hub/api/users")
        if resp.status_code != 200:
            return {}
        payload = resp.json()
        if not isinstance(payload, list):
            return {}
    except Exception:
        return {}

    result = {}
    for item in payload:
        if not isinstance(item, dict):
            continue
        name = _normalize_text(item.get("name"))
        if not name:
            continue
        result[name] = item
    return result


def _extract_server_state(hub_user: Optional[dict]) -> dict:
    payload = hub_user or {}
    pending = bool(payload.get("pending"))
    last_activity = payload.get("last_activity")
    running = False
    server_url = ""

    servers = payload.get("servers")
    if isinstance(servers, dict):
        for _, item in servers.items():
            if not item:
                continue
            if isinstance(item, dict):
                if item.get("pending"):
                    pending = True
                item_url = _normalize_text(item.get("url"))
                if item_url:
                    running = True
                    server_url = server_url or item_url
                elif item.get("ready"):
                    running = True
            else:
                running = True

    server_field = payload.get("server")
    if server_field:
        running = True
        if isinstance(server_field, str):
            server_url = server_url or server_field

    return {
        "server_running": running,
        "server_pending": pending,
        "server_url": server_url,
        "last_activity": last_activity,
        "hub_admin": bool(payload.get("admin")),
    }


def _collect_resource_control_users(overrides: Optional[dict] = None) -> List[dict]:
    users = _managed_users()
    hub_map = _hub_user_state_map()
    rows = []
    for item in users:
        username = item["username"]
        role = item["role"]
        quota, source, meta = _get_effective_user_quota(username, role, overrides=overrides)
        hub_state = _extract_server_state(hub_map.get(username))
        rows.append({
            **item,
            "quota": quota,
            "quota_source": source,
            "quota_updated_by": meta.get("updated_by", ""),
            "quota_updated_at": meta.get("updated_at", ""),
            "quota_note": meta.get("note", ""),
            **hub_state,
        })
    return rows


def _resource_assignment_summary(rows: List[dict], budget: dict) -> dict:
    assigned_cpu = 0.0
    assigned_memory = 0
    assigned_storage = 0
    active_cpu = 0.0
    active_memory = 0
    active_storage = 0
    running_servers = 0

    for item in rows:
        quota = item.get("quota", {})
        cpu = float(quota.get("cpu_limit", 0.0) or 0.0)
        memory = _size_to_bytes(str(quota.get("memory_limit", "0B")))
        storage = _size_to_bytes(str(quota.get("storage_limit", "0B")))
        assigned_cpu += cpu
        assigned_memory += memory
        assigned_storage += storage
        if item.get("server_running"):
            running_servers += 1
            active_cpu += cpu
            active_memory += memory
            active_storage += storage

    budget_cpu = float(budget.get("max_total_cpu", DEFAULT_SERVER_RESOURCE_BUDGET["max_total_cpu"]))
    budget_memory = _size_to_bytes(str(budget.get("max_total_memory", DEFAULT_SERVER_RESOURCE_BUDGET["max_total_memory"])))
    budget_storage = _size_to_bytes(str(budget.get("max_total_storage", DEFAULT_SERVER_RESOURCE_BUDGET["max_total_storage"])))

    return {
        "total_users": len(rows),
        "teachers": len([item for item in rows if item["role"] == "teacher"]),
        "students": len([item for item in rows if item["role"] == "student"]),
        "admins": len([item for item in rows if item["role"] == "admin"]),
        "running_servers": running_servers,
        "assigned_cpu": round(assigned_cpu, 3),
        "assigned_memory_bytes": assigned_memory,
        "assigned_storage_bytes": assigned_storage,
        "active_cpu": round(active_cpu, 3),
        "active_memory_bytes": active_memory,
        "active_storage_bytes": active_storage,
        "budget_cpu": budget_cpu,
        "budget_memory_bytes": budget_memory,
        "budget_storage_bytes": budget_storage,
        "assigned_cpu_ratio": round((assigned_cpu / budget_cpu) if budget_cpu > 0 else 0.0, 4),
        "assigned_memory_ratio": round((assigned_memory / budget_memory) if budget_memory > 0 else 0.0, 4),
        "assigned_storage_ratio": round((assigned_storage / budget_storage) if budget_storage > 0 else 0.0, 4),
    }


def _validate_budget(summary: dict, budget: dict):
    if not budget.get("enforce_budget"):
        return
    if summary["assigned_cpu"] > summary["budget_cpu"] + 1e-9:
        raise HTTPException(status_code=409, detail="分配失败：CPU总配额超出服务器预算")
    if summary["assigned_memory_bytes"] > summary["budget_memory_bytes"]:
        raise HTTPException(status_code=409, detail="分配失败：内存总配额超出服务器预算")
    if summary["assigned_storage_bytes"] > summary["budget_storage_bytes"]:
        raise HTTPException(status_code=409, detail="分配失败：存储总配额超出服务器预算")


def _infer_admission_year(student_id: str) -> str:
    normalized_student_id = _normalize_text(student_id)
    if len(normalized_student_id) >= 2 and normalized_student_id[:2].isdigit():
        return f"20{normalized_student_id[:2]}"
    return ""


def _format_admission_year_label(admission_year: str) -> str:
    normalized = _normalize_admission_year(admission_year)
    return f"{normalized}级" if normalized else ""


def _build_class_name(admission_year: str, major_name: str, class_name: str) -> str:
    normalized_year = _normalize_admission_year(admission_year)
    normalized_major = _normalize_text(major_name)
    normalized_class = _normalize_text(class_name)
    if not (normalized_year and normalized_major and normalized_class):
        return ""
    return f"{normalized_year}级{normalized_major}{normalized_class}"


def _ensure_teacher(teacher_username: str):
    normalized_teacher = _normalize_text(teacher_username)
    if not (is_teacher(normalized_teacher) or is_admin(normalized_teacher)):
        raise HTTPException(status_code=403, detail="权限不足")


def _is_admin_user(username: str) -> bool:
    return is_admin(_normalize_text(username))


def _list_accessible_classes(teacher_username: str) -> List[ClassRecord]:
    normalized_teacher = _normalize_text(teacher_username)
    if _is_admin_user(normalized_teacher):
        return list(classes_db.values())
    return [
        item
        for item in classes_db.values()
        if _normalize_text(item.created_by) == normalized_teacher
    ]


def _student_owner_username(record: StudentRecord) -> str:
    normalized_owner = _normalize_text(record.created_by)
    if normalized_owner:
        return normalized_owner

    matched_class_owners = {
        _normalize_text(item.created_by)
        for item in classes_db.values()
        if item.name == record.class_name and _normalize_text(item.created_by)
    }
    if len(matched_class_owners) == 1:
        return next(iter(matched_class_owners))
    return ""


def _student_visible_to_teacher(record: StudentRecord, teacher_username: str) -> bool:
    normalized_teacher = _normalize_text(teacher_username)
    if _is_admin_user(normalized_teacher):
        return True
    return _student_owner_username(record) == normalized_teacher


def _ensure_student(student_id: str):
    normalized_student_id = _normalize_text(student_id)
    if not normalized_student_id or normalized_student_id not in students_db:
        raise HTTPException(status_code=404, detail="学生不存在")


def _normalize_publish_scope(value: Optional[str]) -> PublishScope:
    if isinstance(value, PublishScope):
        raw = value.value
    else:
        raw = _normalize_text(value).lower()
    if raw.startswith("publishscope."):
        raw = raw.split(".", 1)[1]
    if raw == PublishScope.CLASS.value:
        return PublishScope.CLASS
    if raw == PublishScope.STUDENT.value:
        return PublishScope.STUDENT
    return PublishScope.ALL


def _normalize_experiment_publish_targets(record: Experiment):
    record.publish_scope = _normalize_publish_scope(getattr(record, "publish_scope", PublishScope.ALL.value))

    normalized_classes: List[str] = []
    class_seen: Set[str] = set()
    for item in list(getattr(record, "target_class_names", []) or []):
        normalized = _normalize_text(item)
        key = normalized.lower()
        if not normalized or key in class_seen:
            continue
        class_seen.add(key)
        normalized_classes.append(normalized)
    record.target_class_names = normalized_classes

    normalized_students: List[str] = []
    student_seen: Set[str] = set()
    for item in list(getattr(record, "target_student_ids", []) or []):
        normalized = _normalize_text(item)
        key = normalized.lower()
        if not normalized or key in student_seen:
            continue
        student_seen.add(key)
        normalized_students.append(normalized)
    record.target_student_ids = normalized_students

    if record.publish_scope == PublishScope.ALL:
        record.target_class_names = []
        record.target_student_ids = []
    elif record.publish_scope == PublishScope.CLASS:
        record.target_student_ids = []
    elif record.publish_scope == PublishScope.STUDENT:
        record.target_class_names = []


def _validate_experiment_publish_targets(record: Experiment):
    if not record.published:
        return

    if record.publish_scope == PublishScope.CLASS and not record.target_class_names:
        raise HTTPException(status_code=400, detail="发布范围为班级时，至少选择一个班级")

    if record.publish_scope == PublishScope.STUDENT and not record.target_student_ids:
        raise HTTPException(status_code=400, detail="发布范围为学生时，至少选择一个学生")


def _is_experiment_visible_to_student(record: Experiment, student: StudentRecord) -> bool:
    if not record.published:
        return False

    _normalize_experiment_publish_targets(record)
    if record.publish_scope == PublishScope.ALL:
        return True

    if record.publish_scope == PublishScope.CLASS:
        normalized_targets = {_normalize_text(name) for name in (record.target_class_names or []) if _normalize_text(name)}
        return _normalize_text(student.class_name) in normalized_targets

    if record.publish_scope == PublishScope.STUDENT:
        normalized_targets = {_normalize_text(item) for item in (record.target_student_ids or []) if _normalize_text(item)}
        return _normalize_text(student.student_id) in normalized_targets

    return False


def _is_known_user(username: str) -> bool:
    normalized = _normalize_text(username)
    if not normalized:
        return False
    return is_teacher(normalized) or is_admin(normalized) or normalized in students_db


def _normalize_ai_shared_config(raw: Optional[dict]) -> dict:
    payload = raw or {}
    chat_model = _normalize_text(payload.get("chat_model")) or DEFAULT_AI_SHARED_CONFIG["chat_model"]
    reasoner_model = _normalize_text(payload.get("reasoner_model")) or DEFAULT_AI_SHARED_CONFIG["reasoner_model"]
    base_url = _normalize_text(payload.get("base_url")) or DEFAULT_AI_SHARED_CONFIG["base_url"]
    system_prompt = _normalize_text(payload.get("system_prompt")) or DEFAULT_AI_SHARED_CONFIG["system_prompt"]
    api_key = _normalize_text(payload.get("api_key"))
    tavily_api_key = _normalize_text(payload.get("tavily_api_key"))

    return {
        "api_key": api_key[:512],
        "tavily_api_key": tavily_api_key[:512],
        "chat_model": chat_model[:120],
        "reasoner_model": reasoner_model[:120],
        "base_url": base_url[:500].rstrip("/") or DEFAULT_AI_SHARED_CONFIG["base_url"],
        "system_prompt": system_prompt[:4000],
    }


def _save_ai_shared_config():
    tmp_path = f"{AI_SHARED_CONFIG_FILE}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as file_obj:
        json.dump(ai_shared_config_db, file_obj, ensure_ascii=False, indent=2)
    os.replace(tmp_path, AI_SHARED_CONFIG_FILE)


def _load_ai_shared_config():
    ai_shared_config_db.clear()
    ai_shared_config_db.update(DEFAULT_AI_SHARED_CONFIG)

    if not os.path.exists(AI_SHARED_CONFIG_FILE):
        return

    try:
        with open(AI_SHARED_CONFIG_FILE, "r", encoding="utf-8") as file_obj:
            payload = json.load(file_obj) or {}
        ai_shared_config_db.update(_normalize_ai_shared_config(payload))
    except Exception as exc:
        print(f"Failed to load ai shared config: {exc}")


def _normalize_chat_history_message(raw: Optional[dict]) -> Optional[Dict[str, str]]:
    if not isinstance(raw, dict):
        return None

    role = _normalize_text(raw.get("role")).lower()
    if role not in {"system", "user", "assistant"}:
        return None

    content = str(raw.get("content") or "").strip()
    if not content:
        return None

    return {
        "role": role,
        "content": content[:AI_CHAT_HISTORY_MAX_MESSAGE_CHARS],
    }


def _normalize_chat_history_items(raw_items) -> List[Dict[str, str]]:
    output: List[Dict[str, str]] = []
    for item in raw_items if isinstance(raw_items, list) else []:
        normalized = _normalize_chat_history_message(item)
        if normalized:
            output.append(normalized)
    return output[-AI_CHAT_HISTORY_MAX_MESSAGES:]


def _save_ai_chat_history():
    payload = {}
    for username, items in ai_chat_history_db.items():
        normalized_username = _normalize_text(username)
        if not normalized_username:
            continue
        payload[normalized_username] = _normalize_chat_history_items(items)

    tmp_path = f"{AI_CHAT_HISTORY_FILE}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as file_obj:
        json.dump(payload, file_obj, ensure_ascii=False, indent=2)
    os.replace(tmp_path, AI_CHAT_HISTORY_FILE)


def _load_ai_chat_history():
    ai_chat_history_db.clear()
    if not os.path.exists(AI_CHAT_HISTORY_FILE):
        return

    try:
        with open(AI_CHAT_HISTORY_FILE, "r", encoding="utf-8") as file_obj:
            payload = json.load(file_obj) or {}
    except Exception as exc:
        print(f"Failed to load ai chat history: {exc}")
        return

    if not isinstance(payload, dict):
        return

    for username, items in payload.items():
        normalized_username = _normalize_text(username)
        if not normalized_username:
            continue
        ai_chat_history_db[normalized_username] = _normalize_chat_history_items(items)


def _get_ai_chat_history(username: str) -> List[Dict[str, str]]:
    normalized_username = _normalize_text(username)
    if not normalized_username:
        return []
    return deepcopy(ai_chat_history_db.get(normalized_username, []))


def _set_ai_chat_history(username: str, raw_items) -> List[Dict[str, str]]:
    normalized_username = _normalize_text(username)
    if not normalized_username:
        return []
    normalized_items = _normalize_chat_history_items(raw_items)
    if normalized_items:
        ai_chat_history_db[normalized_username] = normalized_items
    else:
        ai_chat_history_db.pop(normalized_username, None)
    _save_ai_chat_history()
    return deepcopy(normalized_items)


def _trim_ai_history_for_context(raw_items) -> List[Dict[str, str]]:
    normalized_items = _normalize_chat_history_items(raw_items)
    if len(normalized_items) > AI_CONTEXT_MAX_HISTORY_MESSAGES:
        normalized_items = normalized_items[-AI_CONTEXT_MAX_HISTORY_MESSAGES:]

    total_chars = 0
    selected: List[Dict[str, str]] = []
    for item in reversed(normalized_items):
        content = item.get("content", "")
        estimated_chars = len(content) + 16
        if selected and (total_chars + estimated_chars > AI_CONTEXT_MAX_TOTAL_CHARS):
            break
        selected.append({"role": item.get("role", "user"), "content": content})
        total_chars += estimated_chars

    if not selected and normalized_items:
        last_item = normalized_items[-1]
        selected.append({
            "role": last_item.get("role", "user"),
            "content": str(last_item.get("content") or "")[:AI_CONTEXT_MAX_TOTAL_CHARS],
        })

    return list(reversed(selected))


def _class_to_dict(record: ClassRecord) -> dict:
    payload = record.dict()
    payload["created_at"] = record.created_at.isoformat()
    return payload


def _teacher_to_dict(record: TeacherRecord) -> dict:
    payload = record.dict()
    payload["created_at"] = record.created_at.isoformat()
    return payload


def _student_to_dict(record: StudentRecord) -> dict:
    payload = record.dict()
    payload["created_at"] = record.created_at.isoformat()
    payload["updated_at"] = record.updated_at.isoformat()
    return payload


def _save_user_registry():
    default_hash = _default_password_hash()
    account_password_hashes = {}
    for account, password_hash in teacher_account_password_hashes_db.items():
        normalized_account = _normalize_text(account)
        normalized_hash = _normalize_text(password_hash).lower()
        if not normalized_account:
            continue
        if not (is_teacher(normalized_account) or is_admin(normalized_account)):
            continue
        if not PASSWORD_HASH_PATTERN.fullmatch(normalized_hash):
            continue
        if normalized_hash == default_hash:
            continue
        account_password_hashes[normalized_account] = normalized_hash

    account_security_questions = {}
    for account, payload in account_security_questions_db.items():
        normalized_account = _normalize_text(account)
        if not normalized_account:
            continue
        if not (is_teacher(normalized_account) or is_admin(normalized_account)):
            continue

        raw_question = payload or {}
        normalized_question = _normalize_security_question(raw_question.get("question") or "")
        normalized_answer_hash = _normalize_text(raw_question.get("answer_hash") or "").lower()
        if not normalized_question:
            continue
        if not PASSWORD_HASH_PATTERN.fullmatch(normalized_answer_hash):
            continue

        account_security_questions[normalized_account] = {
            "question": normalized_question,
            "answer_hash": normalized_answer_hash,
        }

    payload = {
        "classes": [_class_to_dict(item) for item in classes_db.values()],
        "teachers": [_teacher_to_dict(item) for item in teachers_db.values()],
        "students": [_student_to_dict(item) for item in students_db.values()],
        "account_password_hashes": account_password_hashes,
        "account_security_questions": account_security_questions,
    }
    tmp_path = f"{USER_REGISTRY_FILE}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as file_obj:
        json.dump(payload, file_obj, ensure_ascii=False, indent=2)
    os.replace(tmp_path, USER_REGISTRY_FILE)


def _load_user_registry():
    classes_db.clear()
    teachers_db.clear()
    students_db.clear()
    teacher_account_password_hashes_db.clear()
    account_security_questions_db.clear()

    if not os.path.exists(USER_REGISTRY_FILE):
        return

    try:
        with open(USER_REGISTRY_FILE, "r", encoding="utf-8") as file_obj:
            payload = json.load(file_obj)
    except Exception as exc:
        print(f"Failed to load user registry: {exc}")
        return

    for item in payload.get("classes", []):
        try:
            record = ClassRecord(**item)
            classes_db[record.id] = record
        except Exception as exc:
            print(f"Invalid class record skipped: {exc}")

    for item in payload.get("teachers", []):
        try:
            record = TeacherRecord(**item)
            normalized_username = _normalize_text(record.username)
            if not normalized_username:
                continue
            record.username = normalized_username
            if not record.real_name:
                record.real_name = normalized_username
            teachers_db[normalized_username] = record
        except Exception as exc:
            print(f"Invalid teacher record skipped: {exc}")

    for item in payload.get("students", []):
        try:
            record = StudentRecord(**item)
            if not record.admission_year:
                record.admission_year = _infer_admission_year(record.student_id)
            if not _normalize_text(record.created_by):
                record.created_by = _student_owner_username(record)
            students_db[record.student_id] = record
        except Exception as exc:
            print(f"Invalid student record skipped: {exc}")

    raw_hashes = payload.get("account_password_hashes")
    if isinstance(raw_hashes, dict):
        for account, password_hash in raw_hashes.items():
            normalized_account = _normalize_text(account)
            normalized_hash = _normalize_text(password_hash).lower()
            if not normalized_account:
                continue
            if not (is_teacher(normalized_account) or is_admin(normalized_account)):
                continue
            if not PASSWORD_HASH_PATTERN.fullmatch(normalized_hash):
                continue
            teacher_account_password_hashes_db[normalized_account] = normalized_hash

    raw_security_questions = payload.get("account_security_questions")
    if isinstance(raw_security_questions, dict):
        for account, item in raw_security_questions.items():
            normalized_account = _normalize_text(account)
            if not normalized_account:
                continue
            if not (is_teacher(normalized_account) or is_admin(normalized_account)):
                continue

            raw_item = item if isinstance(item, dict) else {}
            normalized_question = _normalize_security_question(raw_item.get("question") or "")
            normalized_answer_hash = _normalize_text(raw_item.get("answer_hash") or "").lower()
            if not normalized_question:
                continue
            if not PASSWORD_HASH_PATTERN.fullmatch(normalized_answer_hash):
                continue

            account_security_questions_db[normalized_account] = {
                "question": normalized_question,
                "answer_hash": normalized_answer_hash,
            }


def _resource_to_dict(record: ResourceFile) -> dict:
    payload = record.dict()
    payload["created_at"] = record.created_at.isoformat()
    return payload


def _save_resource_registry():
    payload = {
        "items": [_resource_to_dict(item) for item in resource_files_db.values()],
    }
    tmp_path = f"{RESOURCE_REGISTRY_FILE}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as file_obj:
        json.dump(payload, file_obj, ensure_ascii=False, indent=2)
    os.replace(tmp_path, RESOURCE_REGISTRY_FILE)


def _load_resource_registry():
    resource_files_db.clear()
    if not os.path.exists(RESOURCE_REGISTRY_FILE):
        return

    try:
        with open(RESOURCE_REGISTRY_FILE, "r", encoding="utf-8") as file_obj:
            payload = json.load(file_obj)
    except Exception as exc:
        print(f"Failed to load resource registry: {exc}")
        return

    for item in payload.get("items", []):
        try:
            record = ResourceFile(**item)
            if os.path.exists(record.file_path):
                resource_files_db[record.id] = record
        except Exception as exc:
            print(f"Invalid resource record skipped: {exc}")


def _course_to_dict(record: CourseRecord) -> dict:
    payload = record.dict()
    payload["created_at"] = record.created_at.isoformat()
    payload["updated_at"] = record.updated_at.isoformat()
    return payload


def _save_course_registry():
    payload = {
        "courses": [_course_to_dict(item) for item in courses_db.values()],
    }
    tmp_path = f"{COURSE_REGISTRY_FILE}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as file_obj:
        json.dump(payload, file_obj, ensure_ascii=False, indent=2)
    os.replace(tmp_path, COURSE_REGISTRY_FILE)


def _load_course_registry():
    courses_db.clear()
    if not os.path.exists(COURSE_REGISTRY_FILE):
        return

    try:
        with open(COURSE_REGISTRY_FILE, "r", encoding="utf-8") as file_obj:
            payload = json.load(file_obj)
    except Exception as exc:
        print(f"Failed to load course registry: {exc}")
        return

    for item in payload.get("courses", []):
        try:
            record = CourseRecord(**item)
            courses_db[record.id] = record
        except Exception as exc:
            print(f"Invalid course record skipped: {exc}")


def _resolve_course_name(item: Experiment) -> str:
    explicit = _normalize_text(getattr(item, "course_name", ""))
    if explicit:
        return explicit

    notebook_path = _normalize_text(getattr(item, "notebook_path", ""))
    first_segment = next((seg for seg in notebook_path.split("/") if seg), "")
    if first_segment and first_segment.lower() != "course":
        return first_segment

    return "Python程序设计"


def _find_teacher_course_by_name(teacher_username: str, course_name: str) -> Optional[CourseRecord]:
    normalized_teacher = _normalize_text(teacher_username)
    normalized_name = _normalize_text(course_name).lower()
    if not normalized_teacher or not normalized_name:
        return None

    for item in courses_db.values():
        if _normalize_text(item.created_by) != normalized_teacher:
            continue
        if _normalize_text(item.name).lower() == normalized_name:
            return item
    return None


def _create_course_record(name: str, teacher_username: str, description: str = "") -> CourseRecord:
    now = datetime.now()
    record = CourseRecord(
        id=str(uuid.uuid4()),
        name=_normalize_text(name) or "未命名课程",
        description=_normalize_text(description),
        created_by=_normalize_text(teacher_username),
        created_at=now,
        updated_at=now,
    )
    courses_db[record.id] = record
    return record


def _list_course_experiments(course: CourseRecord) -> List[Experiment]:
    normalized_teacher = _normalize_text(course.created_by)
    normalized_course_id = _normalize_text(course.id)
    normalized_course_name = _normalize_text(course.name).lower()

    return [
        item
        for item in experiments_db.values()
        if _normalize_text(item.created_by) == normalized_teacher
        and (
            _normalize_text(item.course_id) == normalized_course_id
            or (
                not _normalize_text(item.course_id)
                and _resolve_course_name(item).lower() == normalized_course_name
            )
        )
    ]


def _resolve_or_create_teacher_course(
    teacher_username: str,
    course_name: str,
    requested_course_id: Optional[str] = None,
) -> Tuple[CourseRecord, bool]:
    normalized_teacher = _normalize_text(teacher_username)
    normalized_name = _normalize_text(course_name) or "Python程序设计"
    normalized_requested_id = _normalize_text(requested_course_id)

    if normalized_requested_id:
        course = courses_db.get(normalized_requested_id)
        if not course:
            raise HTTPException(status_code=404, detail="课程不存在")
        if _normalize_text(course.created_by) != normalized_teacher:
            raise HTTPException(status_code=403, detail="不能使用其他教师创建的课程")
        return course, False

    existing = _find_teacher_course_by_name(normalized_teacher, normalized_name)
    if existing:
        return existing, False

    return _create_course_record(normalized_name, normalized_teacher), True


def _sync_courses_from_experiments():
    experiments_changed = False
    courses_changed = False
    latest_activity: Dict[str, datetime] = {}

    for exp in experiments_db.values():
        teacher_username = _normalize_text(exp.created_by)
        if not teacher_username:
            continue

        course_name = _resolve_course_name(exp)
        requested_course_id = _normalize_text(exp.course_id)
        course_record: Optional[CourseRecord] = None

        if requested_course_id:
            candidate = courses_db.get(requested_course_id)
            if candidate and _normalize_text(candidate.created_by) == teacher_username:
                course_record = candidate

        if course_record is None:
            course_record = _find_teacher_course_by_name(teacher_username, course_name)

        if course_record is None:
            course_record = _create_course_record(course_name, teacher_username)
            courses_changed = True

        if exp.course_id != course_record.id:
            exp.course_id = course_record.id
            experiments_changed = True

        if _normalize_text(exp.course_name) != _normalize_text(course_record.name):
            exp.course_name = course_record.name
            experiments_changed = True

        ts = exp.created_at or datetime.now()
        previous = latest_activity.get(course_record.id)
        if previous is None or ts > previous:
            latest_activity[course_record.id] = ts

    for course_id, latest_ts in latest_activity.items():
        course = courses_db.get(course_id)
        if not course:
            continue
        if course.updated_at is None or latest_ts > course.updated_at:
            course.updated_at = latest_ts
            courses_changed = True

    if experiments_changed:
        _save_experiment_registry()
    if courses_changed:
        _save_course_registry()


def _experiment_to_dict(record: Experiment) -> dict:
    return jsonable_encoder(record)


def _save_experiment_registry():
    payload = {
        "experiments": [_experiment_to_dict(item) for item in experiments_db.values()],
    }
    tmp_path = f"{EXPERIMENT_REGISTRY_FILE}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as file_obj:
        json.dump(payload, file_obj, ensure_ascii=False, indent=2)
    os.replace(tmp_path, EXPERIMENT_REGISTRY_FILE)


def _load_experiment_registry():
    experiments_db.clear()
    if not os.path.exists(EXPERIMENT_REGISTRY_FILE):
        return

    try:
        with open(EXPERIMENT_REGISTRY_FILE, "r", encoding="utf-8") as file_obj:
            payload = json.load(file_obj)
    except Exception as exc:
        print(f"Failed to load experiment registry: {exc}")
        return

    for item in payload.get("experiments", []):
        try:
            record = Experiment(**item)
            if record.id:
                experiments_db[record.id] = record
        except Exception as exc:
            print(f"Invalid experiment skipped: {exc}")


def _get_experiment_by_notebook_path(notebook_path: str) -> Optional[Experiment]:
    needle = _normalize_text(notebook_path)
    if not needle:
        return None
    for exp in experiments_db.values():
        if _normalize_text(exp.notebook_path) == needle:
            return exp
    return None


def _ensure_default_experiments() -> bool:
    """Idempotently seed default experiments when missing."""
    from datetime import timedelta

    seeds = [
        {
            "title": "Python 基础语法练习",
            "description": "本实验旨在帮助你熟悉 Python 的基本语法，包括变量、数据类型、控制流等。",
            "difficulty": "初级",
            "tags": ["Python", "基础", "语法"],
            "notebook_path": "course/python-basics.ipynb",
            "resources": {"cpu": 1.0, "memory": "1G", "storage": "512M"},
            "deadline": datetime.now() + timedelta(days=7),
            "created_by": "teacher_001",
            "published": True,
        },
        {
            "title": "Pandas 数据分析入门",
            "description": "学习使用 Pandas 库进行基本的数据处理和分析操作，包括 DataFrame 的创建、索引、过滤等。",
            "difficulty": "中级",
            "tags": ["Data Science", "Pandas", "数据分析"],
            "notebook_path": "course/pandas-intro.ipynb",
            "resources": {"cpu": 1.0, "memory": "2G", "storage": "1G"},
            "deadline": datetime.now() + timedelta(days=14),
            "created_by": "teacher_001",
            "published": True,
        },
        {
            "title": "机器学习模型训练实战",
            "description": "使用 Scikit-learn 构建一个简单的分类模型，并在真实数据集上进行训练和评估。",
            "difficulty": "高级",
            "tags": ["Machine Learning", "Scikit-learn", "AI"],
            "notebook_path": "course/ml-training.ipynb",
            "resources": {"cpu": 2.0, "memory": "4G", "storage": "2G"},
            "deadline": datetime.now() + timedelta(days=21),
            "created_by": "teacher_001",
            "published": True,
        },
        # --- Extra lab series used by the repo's bundled notebooks ---
        {
            "title": "实验四：NumPy 数组基础与运算",
            "description": "掌握 NumPy ndarray、索引切片、广播与基础运算，为后续数据分析与可视化打基础。",
            "difficulty": "初级",
            "tags": ["Data Science", "NumPy", "实验四"],
            "notebook_path": "course/numpy-lab4.ipynb",
            "resources": {"cpu": 1.0, "memory": "2G", "storage": "1G"},
            "deadline": datetime.now() + timedelta(days=10),
            "created_by": "teacher_001",
            "published": True,
        },
        {
            "title": "实验五：Matplotlib 数据可视化",
            "description": "学习使用 Matplotlib 进行折线图、柱状图、散点图等常用可视化方法。",
            "difficulty": "中级",
            "tags": ["Data Science", "Matplotlib", "实验五"],
            "notebook_path": "course/matplotlib-lab5.ipynb",
            "resources": {"cpu": 1.0, "memory": "2G", "storage": "1G"},
            "deadline": datetime.now() + timedelta(days=12),
            "created_by": "teacher_001",
            "published": True,
        },
        {
            "title": "实验七：Pandas 数据处理与分析",
            "description": "掌握 DataFrame 基础操作、数据清洗、分组聚合与可视化分析的常用流程。",
            "difficulty": "中级",
            "tags": ["Data Science", "Pandas", "实验七"],
            "notebook_path": "course/pandas-lab7.ipynb",
            "resources": {"cpu": 1.0, "memory": "2G", "storage": "1G"},
            "deadline": datetime.now() + timedelta(days=14),
            "created_by": "teacher_001",
            "published": True,
        },
        {
            "title": "综合实验：自动驾驶视觉入门",
            "description": "综合运用 Python、数据处理与计算机视觉基础，完成一个小型自动驾驶视觉实验。",
            "difficulty": "高级",
            "tags": ["Machine Learning", "CV", "综合实验"],
            "notebook_path": "course/autodrive-vision-lab8.ipynb",
            "resources": {"cpu": 2.0, "memory": "4G", "storage": "2G"},
            "deadline": datetime.now() + timedelta(days=21),
            "created_by": "teacher_001",
            "published": True,
        },
    ]

    created = False
    for seed in seeds:
        notebook_path = seed.get("notebook_path") or ""
        if _get_experiment_by_notebook_path(notebook_path) is not None:
            continue

        exp = Experiment(**seed)
        exp.id = str(uuid.uuid4())
        exp.created_at = datetime.now()
        experiments_db[exp.id] = exp
        created = True
        print(f"[seed] Created experiment: {exp.title} ({exp.notebook_path})")

    if created:
        _save_experiment_registry()

    return created


def _normalize_file_type(file_type: str) -> str:
    if not file_type:
        return ""
    return file_type.lower().strip().lstrip(".")


def _resource_preview_mode(record: ResourceFile) -> str:
    file_type = _normalize_file_type(record.file_type)
    if file_type == "pdf":
        return "pdf"
    if file_type in {"xls", "xlsx"}:
        return "sheet"
    if file_type in {"md", "markdown"}:
        return "markdown"
    if file_type in {"txt", "csv", "json", "py", "log"}:
        return "text"
    if file_type == "docx":
        return "docx"
    return "unsupported"


def _read_text_preview(file_path: str) -> str:
    content = None
    for encoding in ("utf-8-sig", "utf-8", "gbk"):
        try:
            with open(file_path, "r", encoding=encoding) as file_obj:
                content = file_obj.read(TEXT_PREVIEW_CHAR_LIMIT + 1)
            break
        except UnicodeDecodeError:
            continue
    if content is None:
        raise HTTPException(status_code=400, detail="文本文件编码无法识别")
    if len(content) > TEXT_PREVIEW_CHAR_LIMIT:
        return f"{content[:TEXT_PREVIEW_CHAR_LIMIT]}\n\n...（预览内容已截断）"
    return content


def _read_docx_preview(file_path: str) -> str:
    try:
        with zipfile.ZipFile(file_path, "r") as archive:
            xml_bytes = archive.read("word/document.xml")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Word 文档解析失败: {exc}") from exc

    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError as exc:
        raise HTTPException(status_code=400, detail=f"Word 文档内容损坏: {exc}") from exc

    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    lines = []
    for paragraph in root.findall(".//w:p", namespace):
        text_parts = [node.text for node in paragraph.findall(".//w:t", namespace) if node.text]
        if text_parts:
            lines.append("".join(text_parts))

    text = "\n".join(lines)
    if len(text) > TEXT_PREVIEW_CHAR_LIMIT:
        return f"{text[:TEXT_PREVIEW_CHAR_LIMIT]}\n\n...（预览内容已截断）"
    return text


def _resource_to_payload(record: ResourceFile, route_prefix: str = "/api/admin/resources") -> dict:
    normalized_prefix = route_prefix.rstrip("/")
    preview_mode = _resource_preview_mode(record)
    return {
        "id": record.id,
        "filename": record.filename,
        "file_type": record.file_type,
        "content_type": record.content_type,
        "size": record.size,
        "created_at": record.created_at,
        "created_by": record.created_by,
        "preview_mode": preview_mode,
        "previewable": preview_mode != "unsupported",
        "preview_url": f"{normalized_prefix}/{record.id}/preview",
        "download_url": f"{normalized_prefix}/{record.id}/download",
    }


def _get_resource_or_404(resource_id: str) -> ResourceFile:
    record = resource_files_db.get(resource_id)
    if not record:
        raise HTTPException(status_code=404, detail="资源文件不存在")
    return record


def _ensure_resource_file_exists(record: ResourceFile):
    if not os.path.exists(record.file_path):
        resource_files_db.pop(record.id, None)
        _save_resource_registry()
        raise HTTPException(status_code=404, detail="资源文件不存在")


def _list_resource_records(name_filter: str = "", type_filter: str = "") -> List[ResourceFile]:
    normalized_name_filter = (name_filter or "").strip().lower()
    normalized_type_filter = _normalize_file_type(type_filter or "")

    items = []
    for record in resource_files_db.values():
        if normalized_name_filter and normalized_name_filter not in record.filename.lower():
            continue
        if normalized_type_filter and _normalize_file_type(record.file_type) != normalized_type_filter:
            continue
        if not os.path.exists(record.file_path):
            continue
        items.append(record)

    items.sort(key=lambda item: item.created_at, reverse=True)
    return items


def _is_template_header(row_values: List[str]) -> bool:
    normalized = [_normalize_text(value) for value in row_values[:len(TEMPLATE_HEADERS)]]
    return (
        normalized[:len(TEMPLATE_HEADERS)] == TEMPLATE_HEADERS
        or normalized[:len(LEGACY_TEMPLATE_HEADERS)] == LEGACY_TEMPLATE_HEADERS
    )


def _read_rows_from_csv(file_content: bytes) -> List[Tuple[int, List[str]]]:
    try:
        content = file_content.decode("utf-8-sig")
    except UnicodeDecodeError:
        content = file_content.decode("gbk")

    parsed_rows: List[Tuple[int, List[str]]] = []
    column_count = len(TEMPLATE_HEADERS)
    reader = csv.reader(io.StringIO(content))
    for row_index, row in enumerate(reader, start=1):
        values = [_normalize_text(value) for value in row[:column_count]]
        while len(values) < column_count:
            values.append("")
        if row_index == 1 and _is_template_header(values):
            continue
        if not any(values):
            continue
        parsed_rows.append((row_index, values))
    return parsed_rows


def _read_rows_from_xlsx(file_content: bytes) -> List[Tuple[int, List[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl is required for xlsx support") from exc

    parsed_rows: List[Tuple[int, List[str]]] = []
    column_count = len(TEMPLATE_HEADERS)
    workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    sheet = workbook.active
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [_normalize_text(value) for value in list(row)[:column_count]]
        while len(values) < column_count:
            values.append("")
        if row_index == 1 and _is_template_header(values):
            continue
        if not any(values):
            continue
        parsed_rows.append((row_index, values))
    workbook.close()
    return parsed_rows


def _parse_student_import_rows(filename: str, file_content: bytes) -> List[Tuple[int, List[str]]]:
    extension = os.path.splitext((filename or "").lower())[1]
    if extension == ".csv":
        return _read_rows_from_csv(file_content)
    if extension == ".xlsx":
        return _read_rows_from_xlsx(file_content)
    raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .csv 文件")


def _build_csv_template() -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(TEMPLATE_HEADERS)
    return buffer.getvalue().encode("utf-8-sig")


def _build_xlsx_template() -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl is required for xlsx support") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "students"
    sheet.append(TEMPLATE_HEADERS)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _is_class_template_header(row_values: List[str]) -> bool:
    normalized = [_normalize_text(value) for value in row_values[:len(CLASS_TEMPLATE_HEADERS)]]
    return normalized == CLASS_TEMPLATE_HEADERS


def _read_class_rows_from_csv(file_content: bytes) -> List[Tuple[int, List[str]]]:
    try:
        content = file_content.decode("utf-8-sig")
    except UnicodeDecodeError:
        content = file_content.decode("gbk")

    parsed_rows: List[Tuple[int, List[str]]] = []
    column_count = len(CLASS_TEMPLATE_HEADERS)
    reader = csv.reader(io.StringIO(content))
    for row_index, row in enumerate(reader, start=1):
        values = [_normalize_text(value) for value in row[:column_count]]
        while len(values) < column_count:
            values.append("")
        if row_index == 1 and _is_class_template_header(values):
            continue
        if not any(values):
            continue
        parsed_rows.append((row_index, values))
    return parsed_rows


def _read_class_rows_from_xlsx(file_content: bytes) -> List[Tuple[int, List[str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl is required for xlsx support") from exc

    parsed_rows: List[Tuple[int, List[str]]] = []
    column_count = len(CLASS_TEMPLATE_HEADERS)
    workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    sheet = workbook.active
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [_normalize_text(value) for value in list(row)[:column_count]]
        while len(values) < column_count:
            values.append("")
        if row_index == 1 and _is_class_template_header(values):
            continue
        if not any(values):
            continue
        parsed_rows.append((row_index, values))
    workbook.close()
    return parsed_rows


def _parse_class_import_rows(filename: str, file_content: bytes) -> List[Tuple[int, List[str]]]:
    extension = os.path.splitext((filename or "").lower())[1]
    if extension == ".csv":
        return _read_class_rows_from_csv(file_content)
    if extension == ".xlsx":
        return _read_class_rows_from_xlsx(file_content)
    raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .csv 文件")


def _build_class_csv_template() -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(CLASS_TEMPLATE_HEADERS)
    return buffer.getvalue().encode("utf-8-sig")


def _build_class_xlsx_template() -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl is required for xlsx support") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "classes"
    sheet.append(CLASS_TEMPLATE_HEADERS)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _get_submission_pdfs(student_exp_id: str) -> List[StudentSubmissionPDF]:
    items = [
        item for item in submission_pdfs_db.values()
        if item.student_exp_id == student_exp_id
    ]
    items.sort(key=lambda item: item.created_at, reverse=True)
    return items


def _pdf_status(item: StudentSubmissionPDF) -> str:
    if item.reviewed:
        return "已批阅"
    if item.viewed:
        return "已查看"
    return "未查看"


def _pdf_to_payload(item: StudentSubmissionPDF) -> dict:
    return {
        "id": item.id,
        "student_exp_id": item.student_exp_id,
        "experiment_id": item.experiment_id,
        "student_id": item.student_id,
        "filename": item.filename,
        "content_type": item.content_type,
        "size": item.size,
        "created_at": item.created_at,
        "download_url": f"/api/student-submissions/{item.id}/download",
        "viewed": item.viewed,
        "viewed_at": item.viewed_at,
        "viewed_by": item.viewed_by,
        "reviewed": item.reviewed,
        "reviewed_at": item.reviewed_at,
        "reviewed_by": item.reviewed_by,
        "review_status": _pdf_status(item),
        "annotations": [
            {
                "id": ann.id,
                "teacher_username": ann.teacher_username,
                "content": ann.content,
                "created_at": ann.created_at,
            }
            for ann in item.annotations
        ],
    }

# ==================== 初始化数据 ====================
@app.on_event("startup")
async def startup_event():
    """服务启动时初始化数据"""
    _load_user_registry()
    print(f"Loaded user registry: {len(classes_db)} classes, {len(teachers_db)} managed teachers, {len(students_db)} students")
    _load_resource_registry()
    print(f"Loaded resource registry: {len(resource_files_db)} files")
    _load_experiment_registry()
    print(f"Loaded experiment registry: {len(experiments_db)} experiments")
    _load_course_registry()
    print(f"Loaded course registry: {len(courses_db)} courses")
    _load_attachment_registry()
    print(f"Loaded attachment registry: {len(attachments_db)} attachments")
    _load_ai_shared_config()
    print("Loaded ai shared config")
    _load_ai_chat_history()
    print(f"Loaded ai chat history users: {len(ai_chat_history_db)}")
    _load_resource_policy()
    print("Loaded resource policy")
    _load_operation_logs()
    print(f"Loaded operation logs: {len(operation_logs_db)}")
    _sync_courses_from_experiments()

    # One-time seeding for this deployment. After seeding, allow teachers to delete/modify
    # experiments without them reappearing on restart.
    seed_version, seed_payload = _read_seed_marker()
    if seed_version == 0:
        _ensure_default_experiments()
        _ensure_default_attachments()
        try:
            _write_seed_marker(2, {"seeded_at": datetime.now().isoformat()})
        except OSError as exc:
            print(f"Failed to write seed marker: {exc}")
    elif seed_version == 1:
        # Migration: older seed logic could attach multiple Word variants. Clean it once.
        try:
            if _cleanup_seeded_attachments():
                print("[seed] Cleaned duplicate seeded attachments")
            _write_seed_marker(2, {"migrated_from": 1, "migrated_at": datetime.now().isoformat()})
        except OSError as exc:
            print(f"Failed to write seed marker: {exc}")

# ==================== API端点 ====================

@app.get("/")
def root():
    return {"message": "福州理工学院AI编程实践教学平台 API", "version": "1.0.0"}


@app.get("/api/health")
def api_health():
    return {
        "status": "ok",
        "service": "experiment-manager",
        "timestamp": datetime.now().isoformat(),
    }


@app.get("/metrics", response_class=PlainTextResponse)
def metrics():
    lines = [
        "# HELP training_backend_up Backend process up status (1=up).",
        "# TYPE training_backend_up gauge",
        "training_backend_up 1",
        "# HELP training_backend_students_total Total students loaded in memory.",
        "# TYPE training_backend_students_total gauge",
        f"training_backend_students_total {len(students_db)}",
        "# HELP training_backend_classes_total Total classes loaded in memory.",
        "# TYPE training_backend_classes_total gauge",
        f"training_backend_classes_total {len(classes_db)}",
        "# HELP training_backend_courses_total Total courses loaded in memory.",
        "# TYPE training_backend_courses_total gauge",
        f"training_backend_courses_total {len(courses_db)}",
        "# HELP training_backend_experiments_total Total experiments loaded in memory.",
        "# TYPE training_backend_experiments_total gauge",
        f"training_backend_experiments_total {len(experiments_db)}",
        "# HELP training_backend_attachments_total Total attachments loaded in memory.",
        "# TYPE training_backend_attachments_total gauge",
        f"training_backend_attachments_total {len(attachments_db)}",
        "# HELP training_backend_resources_total Total uploaded resources loaded in memory.",
        "# TYPE training_backend_resources_total gauge",
        f"training_backend_resources_total {len(resource_files_db)}",
        "# HELP training_backend_resource_quota_overrides_total Total custom user resource overrides.",
        "# TYPE training_backend_resource_quota_overrides_total gauge",
        f"training_backend_resource_quota_overrides_total {len(resource_policy_db.get('overrides', {}))}",
        "# HELP training_backend_operation_logs_total Total operation logs retained.",
        "# TYPE training_backend_operation_logs_total gauge",
        f"training_backend_operation_logs_total {len(operation_logs_db)}",
    ]
    return "\n".join(lines) + "\n"

# ---------- 工作台：用户信息批量导入与管理 ----------

def _list_admin_teacher_items() -> List[dict]:
    rows: List[dict] = []
    for username in _all_teacher_accounts():
        teacher = teachers_db.get(username)
        is_registry = teacher is not None
        rows.append({
            "username": username,
            "real_name": _normalize_text(getattr(teacher, "real_name", "")) or username,
            "source": "registry" if is_registry else "env",
            "created_by": _normalize_text(getattr(teacher, "created_by", "")) or ("system" if is_registry else "env"),
            "created_at": getattr(teacher, "created_at", None),
        })
    return rows











































# ---------- 实验管理 ----------

# ---------- 学生实验管理 ----------

























# ---------- 教师功能 ----------



# ---------- 教师课程管理 ----------

def _course_to_payload(course: CourseRecord) -> dict:
    experiments = sorted(
        _list_course_experiments(course),
        key=lambda item: item.created_at or datetime.min,
        reverse=True,
    )
    published_count = sum(1 for item in experiments if item.published)
    latest_experiment_at = experiments[0].created_at if experiments else None
    tags = sorted(
        {
            tag
            for item in experiments
            for tag in (item.tags or [])
            if _normalize_text(tag)
        }
    )
    return {
        "id": course.id,
        "name": course.name,
        "description": course.description or "",
        "created_by": course.created_by,
        "created_at": course.created_at,
        "updated_at": course.updated_at,
        "experiment_count": len(experiments),
        "published_count": published_count,
        "latest_experiment_at": latest_experiment_at,
        "tags": tags,
        "experiments": experiments,
    }

















# ==================== 附件管理 ====================

class Attachment(BaseModel):
    id: str
    experiment_id: str
    filename: str
    file_path: str
    content_type: str
    size: int
    created_at: datetime

attachments_db = {}


def _attachment_to_dict(record: Attachment) -> dict:
    return jsonable_encoder(record)


def _save_attachment_registry():
    payload = {
        "items": [
            _attachment_to_dict(item)
            for item in attachments_db.values()
            if os.path.exists(item.file_path)
        ],
    }
    tmp_path = f"{ATTACHMENT_REGISTRY_FILE}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as file_obj:
        json.dump(payload, file_obj, ensure_ascii=False, indent=2)
    os.replace(tmp_path, ATTACHMENT_REGISTRY_FILE)


def _load_attachment_registry():
    attachments_db.clear()
    if not os.path.exists(ATTACHMENT_REGISTRY_FILE):
        return

    try:
        with open(ATTACHMENT_REGISTRY_FILE, "r", encoding="utf-8") as file_obj:
            payload = json.load(file_obj)
    except Exception as exc:
        print(f"Failed to load attachment registry: {exc}")
        return

    for item in payload.get("items", []):
        try:
            record = Attachment(**item)
            if record.id and os.path.exists(record.file_path):
                attachments_db[record.id] = record
        except Exception as exc:
            print(f"Invalid attachment skipped: {exc}")


def _find_resource_by_filename(filename: str) -> Optional[ResourceFile]:
    needle = _normalize_text(filename)
    if not needle:
        return None
    for record in resource_files_db.values():
        if _normalize_text(record.filename) == needle:
            return record
    return None


def _find_latest_upload_path_by_suffix(filename_suffix: str) -> Optional[str]:
    """Find latest file in UPLOAD_DIR with the given suffix (case-insensitive)."""
    suffix = _normalize_text(filename_suffix).lower()
    if not suffix:
        return None

    candidates = []
    try:
        entries = os.listdir(UPLOAD_DIR)
    except OSError:
        return None

    for entry in entries:
        if entry.lower().endswith(suffix):
            full_path = os.path.join(UPLOAD_DIR, entry)
            if os.path.isfile(full_path):
                candidates.append(full_path)

    if not candidates:
        return None

    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]


def _attachment_exists(experiment_id: str, filename: str) -> bool:
    exp_id = _normalize_text(experiment_id)
    needle = _normalize_text(filename)
    if not (exp_id and needle):
        return False
    for item in attachments_db.values():
        if item.experiment_id == exp_id and _normalize_text(item.filename) == needle:
            return True
    return False


def _create_attachment_from_file(
    *,
    experiment_id: str,
    filename: str,
    source_path: str,
    content_type: str = "",
) -> Optional[Attachment]:
    exp_id = _normalize_text(experiment_id)
    fname = os.path.basename(_normalize_text(filename))
    src = _normalize_text(source_path)
    if not (exp_id and fname and src):
        return None
    if not os.path.exists(src):
        return None

    att_id = str(uuid.uuid4())
    safe_filename = fname.replace(" ", "_").replace("/", "_").replace("\\", "_")
    dest_path = os.path.join(UPLOAD_DIR, f"{att_id}_{safe_filename}")
    try:
        shutil.copyfile(src, dest_path)
    except Exception as exc:
        print(f"Failed to copy attachment source {src} -> {dest_path}: {exc}")
        return None

    media_type = content_type or mimetypes.guess_type(fname)[0] or "application/octet-stream"
    size = os.path.getsize(dest_path)
    record = Attachment(
        id=att_id,
        experiment_id=exp_id,
        filename=fname,
        file_path=dest_path,
        content_type=media_type,
        size=size,
        created_at=datetime.now(),
    )
    attachments_db[record.id] = record
    return record


def _any_attachment_exists(experiment_id: str, filenames: List[str]) -> bool:
    for name in filenames:
        if _attachment_exists(experiment_id, name):
            return True
    return False


def _ensure_default_attachments() -> bool:
    """Attach bundled guide/docs/templates to seeded experiments (idempotent)."""
    defaults = [
        {
            "notebook_path": "course/numpy-lab4.ipynb",
            "guides": ["lab4-numpy-guide.pdf", "lab4-numpy-guide.doc"],
            "templates": ["学号+姓名+实验四_202510.doc", "学号+姓名+实验四_202510.docx"],
        },
        {
            "notebook_path": "course/matplotlib-lab5.ipynb",
            "guides": ["lab5-matplotlib-guide.pdf", "lab5-matplotlib-guide.doc"],
            "templates": ["学号+姓名+实验五_202511.doc", "学号+姓名+实验五_202511.docx"],
        },
        {
            "notebook_path": "course/pandas-lab7.ipynb",
            "guides": ["lab7-pandas-guide.pdf", "lab7-pandas-guide.doc"],
            "templates": ["学号+姓名+实验七.doc", "学号+姓名+实验七.docx"],
        },
        {
            "notebook_path": "course/autodrive-vision-lab8.ipynb",
            "guides": ["lab8-autodrive-guide.pdf", "lab8-autodrive-guide.doc"],
            "templates": ["学号+姓名+综合实验.doc", "学号+姓名+综合实验.docx"],
        },
    ]

    created_any = False

    for entry in defaults:
        notebook_path = entry.get("notebook_path") or ""
        exp = _get_experiment_by_notebook_path(notebook_path)
        if exp is None:
            continue

        # Guides: attach at most one (prefer PDF, fallback to Word).
        guide_candidates = entry.get("guides", [])
        if not _any_attachment_exists(exp.id, guide_candidates):
            for guide_name in guide_candidates:
                src_path = _find_latest_upload_path_by_suffix(guide_name)
                if not src_path:
                    continue
                record = _create_attachment_from_file(
                    experiment_id=exp.id,
                    filename=guide_name,
                    source_path=src_path,
                    content_type=mimetypes.guess_type(guide_name)[0] or "",
                )
                if record:
                    created_any = True
                    print(f"[seed] Attached {guide_name} -> {exp.title}")
                    break

        # Templates: prefer .doc when present, fallback to .docx.
        template_names = entry.get("templates", [])
        if not _any_attachment_exists(exp.id, template_names):
            for template_name in template_names:
                resource = _find_resource_by_filename(template_name)
                if resource is None:
                    continue
                record = _create_attachment_from_file(
                    experiment_id=exp.id,
                    filename=resource.filename,
                    source_path=resource.file_path,
                    content_type=resource.content_type or "",
                )
                if record:
                    created_any = True
                    print(f"[seed] Attached {resource.filename} -> {exp.title}")
                    break  # Only attach one template variant (.doc preferred via ordering).

    if created_any:
        _save_attachment_registry()

    return created_any


def _cleanup_seeded_attachments() -> bool:
    """Remove duplicate seeded attachments (e.g. doc+docx templates, pdf+doc guides).

    Keep at most one guide (prefer PDF) and one template (prefer .doc, fallback .docx).
    Only touches known seeded filenames.
    """
    groups = [
        {
            "notebook_path": "course/numpy-lab4.ipynb",
            "guides": ["lab4-numpy-guide.pdf", "lab4-numpy-guide.doc"],
            "templates": ["学号+姓名+实验四_202510.doc", "学号+姓名+实验四_202510.docx"],
        },
        {
            "notebook_path": "course/matplotlib-lab5.ipynb",
            "guides": ["lab5-matplotlib-guide.pdf", "lab5-matplotlib-guide.doc"],
            "templates": ["学号+姓名+实验五_202511.doc", "学号+姓名+实验五_202511.docx"],
        },
        {
            "notebook_path": "course/pandas-lab7.ipynb",
            "guides": ["lab7-pandas-guide.pdf", "lab7-pandas-guide.doc"],
            "templates": ["学号+姓名+实验七.doc", "学号+姓名+实验七.docx"],
        },
        {
            "notebook_path": "course/autodrive-vision-lab8.ipynb",
            "guides": ["lab8-autodrive-guide.pdf", "lab8-autodrive-guide.doc"],
            "templates": ["学号+姓名+综合实验.doc", "学号+姓名+综合实验.docx"],
        },
    ]

    removed_any = False

    def _remove_by_filename(exp_id: str, filename: str):
        nonlocal removed_any
        needle = _normalize_text(filename)
        if not needle:
            return
        to_remove = [
            (att_id, item)
            for att_id, item in attachments_db.items()
            if item.experiment_id == exp_id and _normalize_text(item.filename) == needle
        ]
        for att_id, item in to_remove:
            attachments_db.pop(att_id, None)
            removed_any = True
            if item and os.path.exists(item.file_path):
                try:
                    os.remove(item.file_path)
                except OSError:
                    pass

    for group in groups:
        exp = _get_experiment_by_notebook_path(group.get("notebook_path") or "")
        if exp is None:
            continue

        # Guide: prefer PDF, otherwise keep the first existing in list order.
        guide_candidates = group.get("guides", [])
        guide_keep = None
        for name in guide_candidates:
            if _attachment_exists(exp.id, name):
                guide_keep = name
                break
        if guide_keep:
            for name in guide_candidates:
                if name != guide_keep and _attachment_exists(exp.id, name):
                    _remove_by_filename(exp.id, name)

        # Template: prefer .doc, fallback .docx.
        template_candidates = group.get("templates", [])
        template_keep = None
        for name in template_candidates:
            if _attachment_exists(exp.id, name):
                template_keep = name
                break
        if template_keep:
            for name in template_candidates:
                if name != template_keep and _attachment_exists(exp.id, name):
                    _remove_by_filename(exp.id, name)

    if removed_any:
        _save_attachment_registry()

    return removed_any


def _is_pdf_attachment(attachment: Attachment) -> bool:
    lower_filename = attachment.filename.lower()
    return attachment.content_type == "application/pdf" or lower_filename.endswith(".pdf")


def _is_word_filename(filename: str) -> bool:
    lower_filename = filename.lower()
    return lower_filename.endswith(".docx") or lower_filename.endswith(".doc")


def _find_paired_word_attachment(pdf_attachment: Attachment) -> Optional[Attachment]:
    base_name = os.path.splitext(pdf_attachment.filename)[0]
    candidates: List[Attachment] = []

    for item in attachments_db.values():
        if item.experiment_id != pdf_attachment.experiment_id:
            continue
        if os.path.splitext(item.filename)[0] != base_name:
            continue
        if not _is_word_filename(item.filename):
            continue
        if not os.path.exists(item.file_path):
            continue
        candidates.append(item)

    if not candidates:
        return None

    # Prefer docx when both doc and docx exist.
    candidates.sort(key=lambda item: 0 if item.filename.lower().endswith(".docx") else 1)
    return candidates[0]





# ==================== 资源文件管理 ====================






















# ==================== AI集成接口 ====================

class AISharedConfigResponse(BaseModel):
    api_key: str = ""
    tavily_api_key: str = ""
    chat_model: str = DEFAULT_AI_SHARED_CONFIG["chat_model"]
    reasoner_model: str = DEFAULT_AI_SHARED_CONFIG["reasoner_model"]
    base_url: str = DEFAULT_AI_SHARED_CONFIG["base_url"]
    system_prompt: str = DEFAULT_AI_SHARED_CONFIG["system_prompt"]


class AISharedConfigUpdateRequest(BaseModel):
    teacher_username: str = Field(..., min_length=1, max_length=80)
    api_key: str = Field(default="", max_length=512)
    tavily_api_key: str = Field(default="", max_length=512)
    chat_model: str = Field(default=DEFAULT_AI_SHARED_CONFIG["chat_model"], min_length=1, max_length=120)
    reasoner_model: str = Field(default=DEFAULT_AI_SHARED_CONFIG["reasoner_model"], min_length=1, max_length=120)
    base_url: str = Field(default=DEFAULT_AI_SHARED_CONFIG["base_url"], min_length=1, max_length=500)
    system_prompt: str = Field(default=DEFAULT_AI_SHARED_CONFIG["system_prompt"], min_length=1, max_length=4000)


class AIWebSearchRequest(BaseModel):
    query: str = Field(..., min_length=1, max_length=200)
    limit: int = Field(default=5, ge=1, le=8)


class AIChatWithSearchRequest(BaseModel):
    username: str = Field(..., min_length=1, max_length=120)
    message: str = Field(..., min_length=1, max_length=4000)
    history: Optional[List[Dict]] = None
    model: str = Field(default="", max_length=120)
    use_web_search: bool = True
    auto_web_search: bool = True
    search_limit: int = Field(default=4, ge=1, le=8)


class AIChatHistoryMessage(BaseModel):
    role: str = Field(..., min_length=1, max_length=20)
    content: str = Field(..., min_length=1, max_length=AI_CHAT_HISTORY_MAX_MESSAGE_CHARS)


class AIChatHistoryUpdateRequest(BaseModel):
    username: str = Field(..., min_length=1, max_length=120)
    messages: List[AIChatHistoryMessage] = Field(default_factory=list)


class AIChatHistoryResponse(BaseModel):
    username: str
    message_count: int
    messages: List[AIChatHistoryMessage]


def _build_ai_shared_config_response(include_secrets: bool = False) -> AISharedConfigResponse:
    payload = dict(ai_shared_config_db)
    if not include_secrets:
        payload["api_key"] = ""
        payload["tavily_api_key"] = ""
    return AISharedConfigResponse(**payload)


def _strip_html_tags(value: str) -> str:
    if not value:
        return ""
    text = re.sub(r"<[^>]+>", " ", value)
    text = html.unescape(text)
    return re.sub(r"\s+", " ", text).strip()


def _decode_duckduckgo_redirect(url: str) -> str:
    value = html.unescape(url or "").strip()
    if not value:
        return ""
    parsed = urlparse(value)
    if "duckduckgo.com" in parsed.netloc and parsed.path.startswith("/l/"):
        uddg = parse_qs(parsed.query).get("uddg")
        if uddg and uddg[0]:
            return unquote(uddg[0])
    return value


def _extract_duckduckgo_results(html_text: str, limit: int) -> List[Dict[str, str]]:
    link_pattern = re.compile(
        r'<a[^>]*class="result__a"[^>]*href="(?P<url>[^"]+)"[^>]*>(?P<title>.*?)</a>',
        re.IGNORECASE | re.DOTALL,
    )
    snippet_pattern = re.compile(
        r'class="result__snippet"[^>]*>(?P<snippet>.*?)</',
        re.IGNORECASE | re.DOTALL,
    )

    output: List[Dict[str, str]] = []
    seen_urls = set()
    for link_match in link_pattern.finditer(html_text or ""):
        url = _decode_duckduckgo_redirect(link_match.group("url"))
        if not url or url in seen_urls:
            continue

        title = _strip_html_tags(link_match.group("title"))
        nearby_html = (html_text or "")[link_match.end(): link_match.end() + 2200]
        snippet_match = snippet_pattern.search(nearby_html)
        snippet = _strip_html_tags(snippet_match.group("snippet")) if snippet_match else ""

        output.append({
            "title": title or url,
            "url": url,
            "snippet": snippet[:240],
        })
        seen_urls.add(url)
        if len(output) >= limit:
            break
    return output


def _extract_bing_results(html_text: str, limit: int) -> List[Dict[str, str]]:
    pattern = re.compile(
        r'<li class="b_algo".*?<a href="(?P<url>[^"]+)"[^>]*>(?P<title>.*?)</a>.*?(?:<p>(?P<snippet>.*?)</p>)?',
        re.IGNORECASE | re.DOTALL,
    )

    output: List[Dict[str, str]] = []
    seen_urls = set()
    for match in pattern.finditer(html_text or ""):
        url = html.unescape(match.group("url") or "").strip()
        if not url or url in seen_urls:
            continue
        title = _strip_html_tags(match.group("title"))
        snippet = _strip_html_tags(match.group("snippet") or "")
        output.append({
            "title": title or url,
            "url": url,
            "snippet": snippet[:240],
        })
        seen_urls.add(url)
        if len(output) >= limit:
            break
    return output


def _extract_bing_rss_results(xml_text: str, limit: int) -> List[Dict[str, str]]:
    output: List[Dict[str, str]] = []
    seen_urls = set()

    if not xml_text:
        return output

    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return output

    for item in root.findall(".//item"):
        url = (item.findtext("link") or "").strip()
        if not url or url in seen_urls:
            continue

        title = html.unescape((item.findtext("title") or "").strip())
        snippet = html.unescape((item.findtext("description") or "").strip())
        output.append({
            "title": title or url,
            "url": url,
            "snippet": _strip_html_tags(snippet)[:240],
        })
        seen_urls.add(url)
        if len(output) >= limit:
            break

    return output


def _extract_duckduckgo_instant_results(payload: Dict, limit: int) -> List[Dict[str, str]]:
    output: List[Dict[str, str]] = []
    seen_urls = set()

    def _append_result(title: str, url: str, snippet: str):
        cleaned_url = (url or "").strip()
        if not cleaned_url or cleaned_url in seen_urls:
            return
        output.append({
            "title": (title or cleaned_url).strip(),
            "url": cleaned_url,
            "snippet": (snippet or "").strip()[:240],
        })
        seen_urls.add(cleaned_url)

    abstract = _strip_html_tags(str(payload.get("AbstractText") or ""))
    abstract_url = str(payload.get("AbstractURL") or "").strip()
    abstract_title = str(payload.get("Heading") or "").strip() or abstract_url
    if abstract and abstract_url:
        _append_result(abstract_title, abstract_url, abstract)

    related = payload.get("RelatedTopics") or []
    for item in related:
        if len(output) >= limit:
            break

        if isinstance(item, dict) and isinstance(item.get("Topics"), list):
            topics = item.get("Topics") or []
        else:
            topics = [item]

        for topic in topics:
            if len(output) >= limit:
                break
            if not isinstance(topic, dict):
                continue
            topic_url = str(topic.get("FirstURL") or "").strip()
            topic_text = _strip_html_tags(str(topic.get("Text") or ""))
            topic_title = topic_text.split(" - ", 1)[0] if topic_text else topic_url
            _append_result(topic_title, topic_url, topic_text)

    return output[:limit]


def _resolve_tavily_api_key() -> str:
    shared_key = _normalize_text(ai_shared_config_db.get("tavily_api_key"))
    if shared_key:
        return shared_key[:512]
    return TAVILY_API_KEY[:512]


def _cleanup_ai_web_search_cache(now_ts: Optional[float] = None):
    now_value = float(now_ts if now_ts is not None else time.time())

    expired = [
        key
        for key, item in ai_web_search_cache_db.items()
        if float(item.get("expires_at") or 0.0) <= now_value
    ]
    for key in expired:
        ai_web_search_cache_db.pop(key, None)

    if len(ai_web_search_cache_db) <= AI_WEB_SEARCH_CACHE_MAX_ITEMS:
        return

    sorted_items = sorted(
        ai_web_search_cache_db.items(),
        key=lambda pair: float((pair[1] or {}).get("expires_at") or 0.0),
    )
    overflow = len(sorted_items) - AI_WEB_SEARCH_CACHE_MAX_ITEMS
    for key, _ in sorted_items[:overflow]:
        ai_web_search_cache_db.pop(key, None)


def _build_ai_web_search_cache_key(query: str, limit: int, search_depth: str) -> str:
    normalized_query = _normalize_text(query).lower()
    normalized_depth = "advanced" if _normalize_text(search_depth).lower() == "advanced" else "basic"
    normalized_limit = max(1, min(int(limit or 5), 10))
    return f"{normalized_query}|{normalized_depth}|{normalized_limit}"


def _get_ai_web_search_cache(query: str, limit: int, search_depth: str) -> Optional[Dict]:
    _cleanup_ai_web_search_cache()
    key = _build_ai_web_search_cache_key(query, limit, search_depth)
    payload = ai_web_search_cache_db.get(key) or {}
    cached_data = payload.get("data")
    return deepcopy(cached_data) if isinstance(cached_data, dict) else None


def _set_ai_web_search_cache(query: str, limit: int, search_depth: str, payload: Dict):
    if not isinstance(payload, dict):
        return
    _cleanup_ai_web_search_cache()
    key = _build_ai_web_search_cache_key(query, limit, search_depth)
    ai_web_search_cache_db[key] = {
        "expires_at": time.time() + AI_WEB_SEARCH_CACHE_TTL_SECONDS,
        "data": deepcopy(payload),
    }


def _choose_search_depth(query: str) -> str:
    normalized = (query or "").strip().lower()
    if not normalized:
        return "basic"

    advanced_patterns = [
        r"(深度|深入|详细|系统|全面|综述|对比|研究|报告|分析|多来源|论文)",
        r"(advanced|in depth|deep dive|research|compare|benchmark)",
    ]
    if any(re.search(pattern, normalized) for pattern in advanced_patterns):
        return "advanced"

    # 默认优先 basic，降低延迟。
    return "basic"


def _search_with_tavily(query: str, limit: int, search_depth: str = "basic") -> List[Dict[str, str]]:
    api_key = _resolve_tavily_api_key()
    if not api_key:
        return []
    if TavilyClient is None:
        raise RuntimeError("tavily-python dependency is not installed")

    client = TavilyClient(api_key)
    normalized_depth = "advanced" if _normalize_text(search_depth).lower() == "advanced" else "basic"
    max_results = max(1, min(int(limit or 5), 10 if normalized_depth == "advanced" else 5))
    payload = client.search(
        query=query,
        search_depth=normalized_depth,
        max_results=max_results,
        include_answer=True,
        include_raw_content=False,
    ) or {}

    raw_results = payload.get("results") if isinstance(payload, dict) else []
    output: List[Dict[str, str]] = []
    seen_urls = set()
    for item in raw_results or []:
        if not isinstance(item, dict):
            continue
        url = str(item.get("url") or "").strip()
        if not url or url in seen_urls:
            continue
        title = _strip_html_tags(str(item.get("title") or ""))
        snippet = _strip_html_tags(str(item.get("content") or item.get("snippet") or ""))
        output.append({
            "title": title or url,
            "url": url,
            "snippet": snippet[:500],
        })
        seen_urls.add(url)
        if len(output) >= max_results:
            break
    return output


def _build_web_search_context(results: List[Dict[str, str]]) -> str:
    if not results:
        return ""
    lines: List[str] = []
    for index, item in enumerate(results, start=1):
        title = str(item.get("title") or "").strip() or "Untitled"
        url = str(item.get("url") or "").strip()
        snippet = str(item.get("snippet") or "").strip() or "N/A"
        if not url:
            continue
        lines.append(f"{index}. {title}\nURL: {url}\nSummary: {snippet}")
    if not lines:
        return ""
    return f"[WEB_SEARCH_CONTEXT_START]\n{chr(10).join(lines)}\n[WEB_SEARCH_CONTEXT_END]"


def _run_web_search(query: str, limit: int) -> Dict:
    normalized_query = (query or "").strip()
    if not normalized_query:
        raise HTTPException(status_code=400, detail="query 不能为空")

    safe_limit = max(1, min(int(limit or 5), 8))
    search_depth = _choose_search_depth(normalized_query)
    cache_payload = _get_ai_web_search_cache(normalized_query, safe_limit, search_depth)
    if cache_payload:
        cache_payload["cached"] = True
        return cache_payload

    search_queries = _build_search_queries(normalized_query)
    search_errors: List[str] = []
    results: List[Dict[str, str]] = []
    provider = ""
    resolved_query = ""

    for search_query in search_queries:
        try:
            results = _search_with_tavily(search_query, safe_limit, search_depth)
            if results:
                provider = f"tavily-{search_depth}"
                resolved_query = search_query
                break
        except Exception as exc:
            search_errors.append(f"Tavily [{search_query}]: {exc}")

    for search_query in search_queries:
        if results:
            break
        try:
            ddg_html = _request_search_html("https://duckduckgo.com/html/", data={"q": search_query})
            results = _extract_duckduckgo_results(ddg_html, safe_limit)
            provider = "duckduckgo"
            if results:
                resolved_query = search_query
                break
        except requests.RequestException as exc:
            search_errors.append(f"DuckDuckGo [{search_query}]: {exc}")

    if not results:
        for search_query in search_queries:
            try:
                bing_html = _request_search_html(
                    "https://www.bing.com/search",
                    params={"q": search_query, "ensearch": "1"},
                )
                results = _extract_bing_results(bing_html, safe_limit)
                provider = "bing"
                if results:
                    resolved_query = search_query
                    break
            except requests.RequestException as exc:
                search_errors.append(f"Bing [{search_query}]: {exc}")

    if not results:
        for search_query in search_queries:
            try:
                bing_rss = _request_search_html(
                    "https://www.bing.com/search",
                    params={"q": search_query, "format": "rss"},
                )
                results = _extract_bing_rss_results(bing_rss, safe_limit)
                provider = "bing-rss"
                if results:
                    resolved_query = search_query
                    break
            except requests.RequestException as exc:
                search_errors.append(f"Bing RSS [{search_query}]: {exc}")

    if not results:
        for search_query in search_queries:
            try:
                ddg_response = requests.get(
                    "https://api.duckduckgo.com/",
                    params={
                        "q": search_query,
                        "format": "json",
                        "no_html": "1",
                        "no_redirect": "1",
                        "skip_disambig": "1",
                    },
                    timeout=12,
                    headers={
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                    },
                )
                ddg_response.raise_for_status()
                results = _extract_duckduckgo_instant_results(ddg_response.json(), safe_limit)
                provider = "duckduckgo-instant"
                if results:
                    resolved_query = search_query
                    break
            except (requests.RequestException, ValueError) as exc:
                search_errors.append(f"DuckDuckGo Instant [{search_query}]: {exc}")

    if not results and search_errors:
        raise HTTPException(status_code=502, detail=f"联网搜索不可用：{'; '.join(search_errors)}")

    payload = {
        "query": normalized_query,
        "resolved_query": resolved_query or normalized_query,
        "provider": provider or "none",
        "search_depth": search_depth,
        "cached": False,
        "count": len(results),
        "results": results,
    }
    _set_ai_web_search_cache(normalized_query, safe_limit, search_depth, payload)
    return payload


def _chat_completions_url(base_url: str) -> str:
    normalized = _normalize_text(base_url).rstrip("/")
    if not normalized:
        normalized = DEFAULT_AI_SHARED_CONFIG["base_url"].rstrip("/")
    if normalized.endswith("/chat/completions"):
        return normalized
    if normalized.endswith("/v1"):
        return f"{normalized}/chat/completions"
    return f"{normalized}/v1/chat/completions"


def _call_ai_chat_model(*, model: str, messages: List[Dict], base_url: str, api_key: str) -> str:
    if not api_key:
        raise HTTPException(status_code=400, detail="AI API Key 未配置")

    url = _chat_completions_url(base_url)
    try:
        response = requests.post(
            url,
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}",
            },
            json={
                "model": model,
                "stream": False,
                "messages": messages,
            },
            timeout=70,
        )
    except requests.RequestException as exc:
        raise HTTPException(status_code=502, detail=f"调用大模型失败：{exc}") from exc

    raw_text = response.text or ""
    try:
        payload = response.json() if raw_text else {}
    except ValueError:
        payload = {}

    if not response.ok:
        detail = ""
        if isinstance(payload, dict):
            detail = (
                str(((payload.get("error") or {}) if isinstance(payload.get("error"), dict) else {}).get("message") or "")
                or str(payload.get("message") or "")
            )
        detail = detail or raw_text[:300] or f"HTTP {response.status_code}"
        raise HTTPException(status_code=502, detail=f"大模型接口返回错误：{detail}")

    answer = ""
    if isinstance(payload, dict):
        choices = payload.get("choices")
        if isinstance(choices, list) and choices:
            first = choices[0] if isinstance(choices[0], dict) else {}
            message = first.get("message") if isinstance(first, dict) else {}
            if isinstance(message, dict):
                answer = str(message.get("content") or "").strip()

    if not answer:
        raise HTTPException(status_code=502, detail="大模型未返回有效内容")
    return answer


def _extract_json_object(text: str) -> Dict:
    raw = str(text or "").strip()
    if not raw:
        return {}
    try:
        payload = json.loads(raw)
        return payload if isinstance(payload, dict) else {}
    except ValueError:
        pass

    match = re.search(r"\{[\s\S]*\}", raw)
    if not match:
        return {}
    try:
        payload = json.loads(match.group(0))
        return payload if isinstance(payload, dict) else {}
    except ValueError:
        return {}


def _fallback_need_web_search_decision(message: str) -> Tuple[bool, str]:
    text = (message or "").strip().lower()
    if not text:
        return False, "空问题默认不联网"
    patterns = [
        r"(今天|现在|当前|最新|最近|实时|近期|新闻|价格|汇率|天气|股价|比分|赛程|票房|发布日期|官网)",
        r"(what\s+time|what\s+date|latest|news|price|weather|today|current|update)",
    ]
    for pattern in patterns:
        if re.search(pattern, text):
            return True, "规则判断为时效性问题"
    return False, "规则判断为常识/离线可答问题"


def _decide_need_web_search(*, message: str, model: str, base_url: str, api_key: str) -> Tuple[bool, str]:
    decision_system_prompt = (
        "你是联网搜索路由器。只判断当前问题是否需要联网搜索。"
        "当问题涉及时效性信息、最新数据、新闻、价格、天气、日期时间、官网动态时，need_web_search=true。"
        "纯编程解释、数学推导、通用概念、与时效无关内容时，need_web_search=false。"
        "必须只输出 JSON，不要输出其它文本，格式为："
        "{\"need_web_search\": true/false, \"reason\": \"<=30字\"}"
    )
    decision_messages = [
        {"role": "system", "content": decision_system_prompt},
        {"role": "user", "content": message},
    ]
    decision_answer = _call_ai_chat_model(
        model=model,
        messages=decision_messages,
        base_url=base_url,
        api_key=api_key,
    )
    payload = _extract_json_object(decision_answer)
    if not payload:
        return _fallback_need_web_search_decision(message)

    need_value = payload.get("need_web_search")
    if isinstance(need_value, bool):
        need_web_search = need_value
    elif isinstance(need_value, str):
        need_web_search = need_value.strip().lower() in {"1", "true", "yes", "y"}
    else:
        need_web_search, _ = _fallback_need_web_search_decision(message)

    reason = _normalize_text(payload.get("reason"))[:60]
    if not reason:
        reason = "AI已完成联网判定"
    return need_web_search, reason


def _request_search_html(url: str, *, params: Optional[dict] = None, data: Optional[dict] = None) -> str:
    method = "POST" if data is not None else "GET"
    response = requests.request(
        method,
        url,
        params=params,
        data=data,
        timeout=12,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        },
    )
    response.raise_for_status()
    response.encoding = response.encoding or "utf-8"
    return response.text


def _is_datetime_query(query: str) -> bool:
    normalized = (query or "").strip().lower()
    if not normalized:
        return False
    patterns = [
        r"(今天|现在|当前).*(几号|日期|时间|几点|星期)",
        r"(几号|日期|时间|几点|星期).*(今天|现在|当前)",
        r"(what\s+date|what\s+time|current\s+date|current\s+time|today's\s+date)",
        r"(北京时间|上海时间|中国时间|china\s+time|beijing\s+time)",
    ]
    return any(re.search(pattern, normalized) for pattern in patterns)


def _is_today_relative_query(query: str) -> bool:
    normalized = (query or "").strip().lower()
    if not normalized:
        return False
    patterns = [
        r"(今天|今日).*(发生了什么|发生什么|有什么|新闻|热点|头条|消息|动态|事件)",
        r"(发生了什么|发生什么|有什么新闻).*(今天|今日)",
        r"(today\s+news|what\s+happened\s+today|news\s+today)",
    ]
    return any(re.search(pattern, normalized) for pattern in patterns)


def _is_time_sensitive_query(query: str) -> bool:
    normalized = (query or "").strip().lower()
    if not normalized:
        return False
    patterns = [
        r"(今天|今日|现在|当前|最新|最近|实时|近期|刚刚|目前|最新消息|动态|新闻|热点|发生了什么)",
        r"(today|now|current|latest|recent|real[-\s]?time|breaking|news|updates)",
    ]
    return any(re.search(pattern, normalized) for pattern in patterns)


def _current_local_date_tokens() -> Dict[str, str]:
    now_local = datetime.now().astimezone()
    weekday_map = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    weekday = weekday_map[now_local.weekday()]
    return {
        "iso": now_local.strftime("%Y-%m-%d"),
        "cn": f"{now_local.year}年{now_local.month}月{now_local.day}日",
        "compact": now_local.strftime("%Y%m%d"),
        "weekday": weekday,
    }


def _build_search_queries(query: str) -> List[str]:
    base = (query or "").strip()
    if not base:
        return []

    if _is_today_relative_query(base):
        date_tokens = _current_local_date_tokens()
        candidates = [
            f"{base} {date_tokens['cn']}",
            f"{base} {date_tokens['iso']}",
            f"{date_tokens['cn']} {date_tokens['weekday']} 中国 新闻 热点",
            f"{date_tokens['iso']} China today news",
            base,
        ]
        output: List[str] = []
        for item in candidates:
            value = (item or "").strip()
            if value and value not in output:
                output.append(value)
        return output

    if _is_datetime_query(base):
        candidates = [
            f"{base} 实时日期 时间 星期",
            "中国 当前 日期 时间 北京时间 星期几",
            base,
        ]
        output: List[str] = []
        for item in candidates:
            value = (item or "").strip()
            if value and value not in output:
                output.append(value)
        return output

    if _is_time_sensitive_query(base):
        date_tokens = _current_local_date_tokens()
        year = date_tokens["iso"][:4]
        candidates = [
            f"{base} {year}",
            f"{base} 最新动态",
            base,
        ]
        output: List[str] = []
        for item in candidates:
            value = (item or "").strip()
            if value and value not in output:
                output.append(value)
        return output

    return [base]


def _request_network_time(url: str) -> Dict[str, str]:
    response = requests.get(
        url,
        timeout=8,
        allow_redirects=True,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Cache-Control": "no-cache",
        },
    )
    response.raise_for_status()
    date_header = (response.headers.get("Date") or "").strip()
    if not date_header:
        raise ValueError("response missing Date header")

    parsed = parsedate_to_datetime(date_header)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)

    local_time = parsed.astimezone()
    return {
        "source": url,
        "http_date": date_header,
        "utc_iso": parsed.astimezone(timezone.utc).isoformat(),
        "local_iso": local_time.isoformat(),
        "local_readable": local_time.strftime("%Y-%m-%d %H:%M:%S %Z"),
    }


def _fetch_network_time() -> Tuple[Optional[Dict[str, str]], List[str]]:
    providers = [
        "https://www.bing.com/",
        "https://www.baidu.com/",
        "https://www.cloudflare.com/",
    ]
    errors: List[str] = []
    for provider_url in providers:
        try:
            return _request_network_time(provider_url), errors
        except (requests.RequestException, ValueError) as exc:
            errors.append(f"{provider_url}: {exc}")
    return None, errors






















from .api.v1.router import router as api_v1_router

app.include_router(api_v1_router, prefix="")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)








